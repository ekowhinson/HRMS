"""
Report export utilities for generating Excel, CSV, and PDF files.
"""

import csv
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Count, F, Avg

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from employees.models import Employee, EmploymentHistory
from payroll.models import PayrollRun, PayrollItem, PayrollItemDetail
from leave.models import LeaveBalance, LeaveRequest
from benefits.models import LoanAccount
from performance.models import Appraisal
from training.models import TrainingProgram, TrainingSession

from openpyxl.chart import PieChart as XlPieChart, BarChart as XlBarChart, LineChart as XlLineChart, AreaChart as XlAreaChart, Reference as XlReference

from reports.chart_utils import generate_pie_chart, generate_bar_chart, generate_line_chart, generate_area_chart


def decimal_to_float(obj):
    """Convert Decimal to float for serialization."""
    if isinstance(obj, Decimal):
        return float(obj)
    return obj


# =============================================================================
# CSV Export Functions
# =============================================================================

def generate_csv_response(data: list, headers: list, filename: str) -> HttpResponse:
    """Generate a CSV file response."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(headers)

    for row in data:
        row_data = []
        for h in headers:
            key = h.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
            row_data.append(decimal_to_float(row.get(key, '')))
        writer.writerow(row_data)

    return response


# =============================================================================
# Excel Export Functions
# =============================================================================

def generate_excel_response(data: list, headers: list, filename: str, title: str = None) -> HttpResponse:
    """Generate an Excel file response."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 1

    # Add title if provided
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        start_row = 3

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(data, start_row + 1):
        for col_idx, header in enumerate(headers, 1):
            key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
            value = decimal_to_float(row.get(key, ''))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                if 'salary' in key or 'amount' in key or 'balance' in key or 'paye' in key or 'ssnit' in key or 'deduction' in key or 'earning' in key or 'net' in key or 'gross' in key:
                    cell.number_format = '#,##0.00'

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(start_row, len(data) + start_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# PDF Export Functions
# =============================================================================

def generate_pdf_response(data: list, headers: list, filename: str, title: str = None, landscape_mode: bool = False) -> HttpResponse:
    """Generate a PDF file response with proper text wrapping."""
    buffer = io.BytesIO()

    page_size = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    if title:
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))

    # Subtitle with date
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    # Cell styles for text wrapping
    header_cell_style = ParagraphStyle(
        'HeaderCell', parent=styles['Normal'], fontSize=8,
        leading=10, fontName='Helvetica-Bold', alignment=TA_CENTER,
        textColor=colors.whitesmoke, wordWrap='CJK'
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontSize=8,
        leading=10, wordWrap='CJK'
    )
    cell_style_right = ParagraphStyle(
        'CellRight', parent=styles['Normal'], fontSize=8,
        leading=10, wordWrap='CJK', alignment=TA_RIGHT
    )

    # Prepare table data with Paragraph objects for text wrapping
    header_row = [Paragraph(str(h), header_cell_style) for h in headers]
    table_data = [header_row]

    for row in data:
        row_data = []
        for header in headers:
            key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
            value = decimal_to_float(row.get(key, ''))
            if isinstance(value, float):
                # Use right-aligned style for numbers
                row_data.append(Paragraph(f"{value:,.2f}", cell_style_right))
            else:
                row_data.append(Paragraph(str(value) if value else '', cell_style))
        table_data.append(row_data)

    # Calculate column widths based on content
    available_width = page_size[0] - 1*inch
    col_widths = [available_width / len(headers)] * len(headers)

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style - simplified since Paragraph handles text styling
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# PDF/Excel with Charts Export Functions
# =============================================================================

def _render_chart_image(section):
    """Render a chart section dict to a BytesIO PNG image using chart_utils."""
    chart_type = section.get('chart_type')
    data = section.get('data', [])
    title = section.get('title', '')
    options = section.get('options', {})

    if not data:
        return None

    if chart_type == 'pie':
        return generate_pie_chart(data, title, donut=options.get('donut', False),
                                  center_label=options.get('center_label'))
    elif chart_type == 'bar':
        return generate_bar_chart(data, title, horizontal=options.get('horizontal', False),
                                  color=options.get('color'))
    elif chart_type == 'line':
        return generate_line_chart(data, title)
    elif chart_type == 'area':
        return generate_area_chart(data, title, fill_color=options.get('fill_color'))
    return None


def generate_pdf_with_charts_response(sections: list, filename: str, title: str = None,
                                       landscape_mode: bool = False) -> HttpResponse:
    """
    Generate a PDF with embedded chart images and data tables.

    sections: list of dicts, each one of:
        {'type': 'table', 'data': [...], 'headers': [...], 'subtitle': '...'}
        {'type': 'chart', 'chart_type': 'pie|bar|line|area', 'data': [...], 'title': '...', 'options': {...}}
        {'type': 'stats', 'items': [{'label': '...', 'value': '...'}]}
    """
    buffer = io.BytesIO()

    page_size = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    if title:
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=16, alignment=TA_CENTER, spaceAfter=10
        )
        elements.append(Paragraph(title, title_style))

    # Date
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=10, alignment=TA_CENTER, spaceAfter=10
    )
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Section styles
    section_title_style = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'],
        fontSize=12, spaceAfter=8, spaceBefore=12
    )
    header_cell_style = ParagraphStyle(
        'HeaderCell', parent=styles['Normal'], fontSize=8,
        leading=10, fontName='Helvetica-Bold', alignment=TA_CENTER,
        textColor=colors.whitesmoke, wordWrap='CJK'
    )
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontSize=8,
        leading=10, wordWrap='CJK'
    )
    cell_style_right = ParagraphStyle(
        'CellRight', parent=styles['Normal'], fontSize=8,
        leading=10, wordWrap='CJK', alignment=TA_RIGHT
    )
    stats_label_style = ParagraphStyle(
        'StatsLabel', parent=styles['Normal'], fontSize=10,
        leading=14, fontName='Helvetica-Bold'
    )
    stats_value_style = ParagraphStyle(
        'StatsValue', parent=styles['Normal'], fontSize=10,
        leading=14, alignment=TA_RIGHT
    )

    for section in sections:
        if section['type'] == 'chart':
            chart_image = _render_chart_image(section)
            if chart_image:
                img = RLImage(chart_image, width=5 * inch, height=3.5 * inch)
                elements.append(img)
                elements.append(Spacer(1, 0.3 * inch))

        elif section['type'] == 'table':
            if section.get('subtitle'):
                elements.append(Paragraph(section['subtitle'], section_title_style))

            table_headers = section['headers']
            table_data = section['data']

            header_row = [Paragraph(str(h), header_cell_style) for h in table_headers]
            rows = [header_row]

            for row in table_data:
                row_cells = []
                for header in table_headers:
                    key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
                    value = decimal_to_float(row.get(key, ''))
                    if isinstance(value, float):
                        row_cells.append(Paragraph(f"{value:,.2f}", cell_style_right))
                    else:
                        row_cells.append(Paragraph(str(value) if value else '', cell_style))
                rows.append(row_cells)

            available_width = page_size[0] - 1 * inch
            col_widths = [available_width / len(table_headers)] * len(table_headers)

            table = Table(rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))

        elif section['type'] == 'stats':
            if section.get('subtitle'):
                elements.append(Paragraph(section['subtitle'], section_title_style))
            stat_rows = []
            for item in section['items']:
                stat_rows.append([
                    Paragraph(str(item['label']), stats_label_style),
                    Paragraph(str(item['value']), stats_value_style),
                ])
            if stat_rows:
                stat_table = Table(stat_rows, colWidths=[3.5 * inch, 2 * inch])
                stat_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F8F8')),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(stat_table)
                elements.append(Spacer(1, 0.3 * inch))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_excel_with_charts_response(sections: list, filename: str, title: str = None) -> HttpResponse:
    """
    Generate an Excel file with native charts and data tables.

    sections: list of dicts, each one of:
        {'type': 'table', 'data': [...], 'headers': [...], 'subtitle': '...'}
        {'type': 'chart', 'chart_type': 'pie|bar|line|area', 'data': [...], 'title': '...', 'options': {...}}
    """
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Data sheet ---
    ws = wb.active
    ws.title = "Report"
    current_row = 1

    if title:
        # Find the max column count from table sections
        max_cols = 2
        for s in sections:
            if s['type'] == 'table' and s.get('headers'):
                max_cols = max(max_cols, len(s['headers']))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        current_row = 3

    for section in sections:
        if section['type'] != 'table':
            continue

        table_headers = section.get('headers', [])
        table_data = section.get('data', [])
        subtitle = section.get('subtitle')

        if subtitle:
            ws.cell(row=current_row, column=1, value=subtitle).font = Font(bold=True, size=11)
            current_row += 1

        # Write headers
        for col, header in enumerate(table_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(table_data, current_row + 1):
            for col_idx, header in enumerate(table_headers, 1):
                key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
                value = decimal_to_float(row.get(key, ''))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    if any(k in key for k in ['salary', 'amount', 'balance', 'paye', 'ssnit',
                                               'deduction', 'earning', 'net', 'gross']):
                        cell.number_format = '#,##0.00'

        current_row += len(table_data) + 3

        # Auto-adjust column widths
        for col in range(1, len(table_headers) + 1):
            max_length = len(str(table_headers[col - 1]))
            column_letter = get_column_letter(col)
            for r in range(current_row - len(table_data) - 2, current_row - 2):
                cell = ws.cell(row=r, column=col)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # --- Charts sheet ---
    chart_sections = [s for s in sections if s['type'] == 'chart' and s.get('data')]
    if chart_sections:
        cs = wb.create_sheet("Charts")
        chart_data_row = 1
        chart_placement_col = 5  # Column E - where charts are placed
        chart_y_offset = 0

        for section in chart_sections:
            chart_data = section['data']
            chart_type = section['chart_type']
            chart_title = section.get('title', '')
            options = section.get('options', {})

            if not chart_data:
                continue

            # Write chart source data
            cs.cell(row=chart_data_row, column=1, value='Name').font = Font(bold=True)
            cs.cell(row=chart_data_row, column=2, value='Value').font = Font(bold=True)
            data_start = chart_data_row + 1
            for i, item in enumerate(chart_data):
                cs.cell(row=data_start + i, column=1, value=str(item.get('name', '')))
                cs.cell(row=data_start + i, column=2, value=item.get('value', 0))
            data_end = data_start + len(chart_data) - 1

            # Auto-width for label column
            max_label_len = max((len(str(item.get('name', ''))) for item in chart_data), default=10)
            cs.column_dimensions['A'].width = max(cs.column_dimensions['A'].width or 10, min(max_label_len + 2, 40))
            cs.column_dimensions['B'].width = 12

            # Create native chart
            chart = None
            if chart_type == 'pie':
                chart = XlPieChart()
                chart.style = 10
                data_ref = XlReference(cs, min_col=2, min_row=chart_data_row, max_row=data_end)
                cats = XlReference(cs, min_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
            elif chart_type == 'bar':
                chart = XlBarChart()
                if options.get('horizontal'):
                    chart.type = "bar"
                chart.style = 10
                data_ref = XlReference(cs, min_col=2, min_row=chart_data_row, max_row=data_end)
                cats = XlReference(cs, min_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.legend = None
            elif chart_type == 'line':
                chart = XlLineChart()
                chart.style = 10
                data_ref = XlReference(cs, min_col=2, min_row=chart_data_row, max_row=data_end)
                cats = XlReference(cs, min_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.legend = None
            elif chart_type == 'area':
                chart = XlAreaChart()
                chart.style = 10
                data_ref = XlReference(cs, min_col=2, min_row=chart_data_row, max_row=data_end)
                cats = XlReference(cs, min_col=1, min_row=data_start, max_row=data_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.legend = None

            if chart:
                chart.title = chart_title
                chart.width = 18
                chart.height = 12
                # Place charts in column E, stacked vertically
                placement_cell = f"E{chart_data_row}"
                cs.add_chart(chart, placement_cell)

            chart_data_row = data_end + 3

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# Report Data Extraction Functions
# =============================================================================

def apply_employee_filters(queryset, filters: dict):
    """Apply common employee filters to a queryset."""
    if not filters:
        return queryset

    if filters.get('employee_code'):
        queryset = queryset.filter(employee_number__icontains=filters['employee_code'])
    if filters.get('division'):
        queryset = queryset.filter(division_id=filters['division'])
    if filters.get('directorate'):
        queryset = queryset.filter(directorate_id=filters['directorate'])
    if filters.get('department'):
        queryset = queryset.filter(department_id=filters['department'])
    if filters.get('position'):
        queryset = queryset.filter(position_id=filters['position'])
    if filters.get('grade'):
        queryset = queryset.filter(grade_id=filters['grade'])
    if filters.get('salary_band'):
        # Salary band is accessed via salary_notch -> level -> band
        queryset = queryset.filter(salary_notch__level__band_id=filters['salary_band'])
    if filters.get('salary_level'):
        # Salary level is accessed via salary_notch -> level
        queryset = queryset.filter(salary_notch__level_id=filters['salary_level'])
    if filters.get('staff_category'):
        queryset = queryset.filter(staff_category_id=filters['staff_category'])
    if filters.get('status'):
        queryset = queryset.filter(status=filters['status'].upper())

    return queryset


def get_employee_master_data(filters: dict = None):
    """Get employee master report data."""
    queryset = Employee.objects.select_related(
        'department', 'grade', 'position', 'division', 'directorate',
        'salary_notch', 'salary_notch__level', 'salary_notch__level__band',
        'staff_category'
    )

    # Default to active if no status filter specified
    if not filters or not filters.get('status'):
        queryset = queryset.filter(status='ACTIVE')

    # Apply all filters
    queryset = apply_employee_filters(queryset, filters)

    headers = [
        'Employee Number', 'First Name', 'Last Name', 'Email', 'Phone',
        'Division', 'Directorate', 'Department', 'Position', 'Grade',
        'Salary Band', 'Salary Level', 'Staff Category', 'Employment Type',
        'Date of Joining', 'Status'
    ]

    data = []
    for emp in queryset:
        # Get salary band and level via salary_notch relationship
        salary_band = ''
        salary_level = ''
        if emp.salary_notch:
            if emp.salary_notch.level:
                salary_level = emp.salary_notch.level.name
                if emp.salary_notch.level.band:
                    salary_band = emp.salary_notch.level.band.name

        data.append({
            'employee_number': emp.employee_number,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'email': emp.work_email or emp.personal_email or '',
            'phone': emp.mobile_phone or '',
            'division': emp.division.name if emp.division else '',
            'directorate': emp.directorate.name if emp.directorate else '',
            'department': emp.department.name if emp.department else '',
            'position': emp.position.title if emp.position else '',
            'grade': emp.grade.name if emp.grade else '',
            'salary_band': salary_band,
            'salary_level': salary_level,
            'staff_category': emp.staff_category.name if emp.staff_category else '',
            'employment_type': emp.employment_type,
            'date_of_joining': emp.date_of_joining.strftime('%Y-%m-%d') if emp.date_of_joining else '',
            'status': emp.status,
        })

    return data, headers


def get_headcount_data(filters: dict = None):
    """Get headcount report data."""
    queryset = Employee.objects.filter(status='ACTIVE')

    # Apply filters
    queryset = apply_employee_filters(queryset, filters)

    by_department = queryset.values(
        department_name=F('department__name')
    ).annotate(count=Count('id')).order_by('-count')

    headers = ['Department', 'Headcount']
    data = [{'department': row['department_name'] or 'Unassigned', 'headcount': row['count']} for row in by_department]

    return data, headers


def apply_payroll_item_filters(queryset, filters: dict):
    """Apply filters to PayrollItem queryset via employee relation."""
    if not filters:
        return queryset

    if filters.get('employee_code'):
        queryset = queryset.filter(employee__employee_number__icontains=filters['employee_code'])
    if filters.get('division'):
        queryset = queryset.filter(employee__division_id=filters['division'])
    if filters.get('directorate'):
        queryset = queryset.filter(employee__directorate_id=filters['directorate'])
    if filters.get('department'):
        queryset = queryset.filter(employee__department_id=filters['department'])
    if filters.get('position'):
        queryset = queryset.filter(employee__position_id=filters['position'])
    if filters.get('grade'):
        queryset = queryset.filter(employee__grade_id=filters['grade'])
    if filters.get('salary_band'):
        # Salary band is accessed via employee -> salary_notch -> level -> band
        queryset = queryset.filter(employee__salary_notch__level__band_id=filters['salary_band'])
    if filters.get('salary_level'):
        # Salary level is accessed via employee -> salary_notch -> level
        queryset = queryset.filter(employee__salary_notch__level_id=filters['salary_level'])
    if filters.get('staff_category'):
        queryset = queryset.filter(employee__staff_category_id=filters['staff_category'])
    if filters.get('bank'):
        queryset = queryset.filter(bank_name__icontains=filters['bank'])

    return queryset


def get_payroll_summary_data(payroll_run_id: str = None, filters: dict = None):
    """Get payroll summary data."""
    if payroll_run_id:
        try:
            payroll_run = PayrollRun.objects.get(id=payroll_run_id)
        except PayrollRun.DoesNotExist:
            payroll_run = None
    else:
        payroll_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()

    if not payroll_run:
        return [], ['No Data']

    items = PayrollItem.objects.filter(
        payroll_run=payroll_run
    ).select_related('employee', 'employee__department', 'employee__division',
                     'employee__directorate', 'employee__position', 'employee__grade')

    # Apply filters
    items = apply_payroll_item_filters(items, filters)

    headers = [
        'Employee Number', 'Employee Name', 'Department',
        'Basic Salary', 'Gross Earnings', 'SSNIT Employee', 'PAYE Tax',
        'Total Deductions', 'Net Salary', 'Bank', 'Account Number'
    ]

    data = []
    for item in items:
        data.append({
            'employee_number': item.employee.employee_number,
            'employee_name': item.employee.full_name,
            'department': item.employee.department.name if item.employee.department else '',
            'basic_salary': float(item.basic_salary),
            'gross_earnings': float(item.gross_earnings),
            'ssnit_employee': float(item.ssnit_employee),
            'paye_tax': float(item.paye),
            'total_deductions': float(item.total_deductions),
            'net_salary': float(item.net_salary),
            'bank': item.bank_name or '',
            'account_number': item.bank_account_number or '',
        })

    return data, headers


def get_paye_data(payroll_run_id: str = None, filters: dict = None):
    """Get PAYE tax report data."""
    if payroll_run_id:
        items = PayrollItem.objects.filter(payroll_run_id=payroll_run_id)
    else:
        latest_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()
        if latest_run:
            items = PayrollItem.objects.filter(payroll_run=latest_run)
        else:
            items = PayrollItem.objects.none()

    items = items.select_related(
        'employee', 'employee__department', 'employee__division',
        'employee__directorate', 'employee__position', 'employee__grade'
    ).filter(paye__gt=0)

    # Apply filters
    items = apply_payroll_item_filters(items, filters)

    headers = [
        'Employee Number', 'Employee Name', 'TIN Number',
        'Gross Salary', 'Taxable Income', 'PAYE Tax'
    ]

    data = []
    for item in items:
        data.append({
            'employee_number': item.employee.employee_number,
            'employee_name': item.employee.full_name,
            'tin_number': item.employee.tin_number or '',
            'gross_salary': float(item.gross_earnings),
            'taxable_income': float(item.taxable_income),
            'paye_tax': float(item.paye),
        })

    return data, headers


def get_paye_gra_data(payroll_run_id: str = None, filters: dict = None):
    """
    Get PAYE data in GRA (Ghana Revenue Authority) format.
    Matches the official PAYE Monthly Tax Deductions Schedule format.
    """
    if payroll_run_id:
        try:
            payroll_run = PayrollRun.objects.get(id=payroll_run_id)
        except PayrollRun.DoesNotExist:
            payroll_run = None
    else:
        payroll_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()

    if not payroll_run:
        return [], None, None

    items = PayrollItem.objects.filter(
        payroll_run=payroll_run
    ).select_related(
        'employee', 'employee__department', 'employee__position',
        'employee__division', 'employee__directorate', 'employee__grade'
    ).prefetch_related(
        'details', 'details__pay_component'
    ).order_by('employee__last_name', 'employee__first_name')

    # Apply filters
    items = apply_payroll_item_filters(items, filters)

    # Get period info for header
    period = payroll_run.payroll_period
    period_str = f"{period.month:02d}/{period.year}" if period else ""

    # Get organization info
    from django.conf import settings
    payroll_settings = getattr(settings, 'PAYROLL', {})
    org_name = payroll_settings.get('ORGANIZATION_NAME', 'ORGANIZATION')

    # Build data rows
    data = []
    for idx, item in enumerate(items, 1):
        emp = item.employee

        # Get component amounts from details
        cash_allowances = 0
        bonus_income = 0
        tier3_contribution = 0

        for detail in item.details.all():
            comp = detail.pay_component
            amount = float(detail.amount)

            if comp.component_type == 'EARNING' and comp.code != 'BASIC':
                # Classify as allowance or bonus based on component category
                if comp.category == 'BONUS':
                    bonus_income += amount
                else:
                    cash_allowances += amount
            elif comp.code in ['TIER3', 'PF_EMPLOYEE', 'PROVIDENT_FUND']:
                tier3_contribution += amount

        # Calculate totals
        basic_salary = float(item.basic_salary)
        ssnit_employee = float(item.ssnit_employee)  # Social Security Fund (Tier 1)
        total_cash_emolument = basic_salary + cash_allowances + bonus_income

        # Final tax on bonus (simplified - usually 5% on bonus up to 15% of annual basic)
        annual_basic = basic_salary * 12
        max_exempt_bonus = annual_basic * 0.15
        final_tax_on_bonus = 0
        excess_bonus = 0
        if bonus_income > 0:
            if bonus_income <= max_exempt_bonus / 12:  # Monthly equivalent
                final_tax_on_bonus = bonus_income * 0.05  # 5% final tax
            else:
                exempt_portion = max_exempt_bonus / 12
                final_tax_on_bonus = exempt_portion * 0.05
                excess_bonus = bonus_income - exempt_portion

        # Accommodation, vehicle, non-cash benefits (from payroll item if available)
        accommodation = 0
        vehicle_element = 0
        non_cash_benefit = 0

        # Total assessable income
        total_assessable = total_cash_emolument + accommodation + vehicle_element + non_cash_benefit

        # Reliefs
        deductible_reliefs = 0  # Personal relief, marriage allowance, etc.
        total_reliefs = ssnit_employee + tier3_contribution + deductible_reliefs

        # Chargeable income
        chargeable_income = max(0, total_assessable - total_reliefs)

        # Tax deductible (PAYE)
        tax_deductible = float(item.paye)

        # Overtime (if applicable)
        overtime_income = float(item.overtime_earnings) if hasattr(item, 'overtime_earnings') and item.overtime_earnings else 0
        overtime_tax = float(item.overtime_tax) if hasattr(item, 'overtime_tax') and item.overtime_tax else 0

        # Total tax payable to GRA
        total_tax_payable = final_tax_on_bonus + tax_deductible + overtime_tax

        data.append({
            'ser_no': idx,
            'tin_number': emp.tin_number or '0',
            'employee_name': emp.full_name,
            'position': emp.position.title if emp.position else '',
            'residency': 'Residency',  # Default to Residency
            'basic_salary': basic_salary,
            'secondary_employment': 'N',
            'paid_ssnit': 'Y' if ssnit_employee > 0 else 'N',
            'social_security_fund': ssnit_employee,
            'third_tier': tier3_contribution,
            'cash_allowance': cash_allowances,
            'bonus_income': bonus_income,
            'final_tax_on_bonus': final_tax_on_bonus,
            'excess_bonus': excess_bonus,
            'total_cash_emolument': total_cash_emolument,
            'accommodation_element': accommodation,
            'vehicle_element': vehicle_element,
            'non_cash_benefit': non_cash_benefit,
            'total_assessable_income': total_assessable,
            'deductible_reliefs': deductible_reliefs,
            'total_reliefs': total_reliefs,
            'chargeable_income': chargeable_income,
            'tax_deductible': tax_deductible,
            'overtime_income': overtime_income,
            'overtime_tax': overtime_tax,
            'total_tax_payable': total_tax_payable,
            'severance_pay': 0,
            'remarks': '',
        })

    return data, org_name, period_str


def export_paye_gra_excel(data: list, org_name: str, period: str, filename: str) -> HttpResponse:
    """
    Generate GRA-compliant PAYE Excel report matching the official format.
    Includes header section with GRA branding and organization details.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PAYE Monthly Schedule"

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    subheader_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # GRA Header Section (Rows 1-6)
    # Row 1: Ghana Revenue Authority
    ws.merge_cells('A1:AB1')
    ws['A1'] = "GHANA REVENUE AUTHORITY"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align

    # Row 2: Domestic Tax Revenue Division
    ws.merge_cells('A2:AB2')
    ws['A2'] = "DOMESTIC TAX REVENUE DIVISION"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center_align

    # Row 3: PAYE Monthly Tax Deductions Schedule
    ws.merge_cells('A3:AB3')
    ws['A3'] = "PAYE MONTHLY TAX DEDUCTIONS SCHEDULE"
    ws['A3'].font = title_font
    ws['A3'].alignment = center_align

    # Row 4: Blank

    # Row 5: Organization Name
    ws.merge_cells('A5:AB5')
    ws['A5'] = org_name
    ws['A5'].font = Font(bold=True, size=12)
    ws['A5'].alignment = center_align

    # Row 6: Period
    ws.merge_cells('A6:AB6')
    ws['A6'] = f"Period: {period}" if period else "Period: N/A"
    ws['A6'].font = header_font
    ws['A6'].alignment = center_align

    # Row 7: Blank

    # Column Headers (Row 8)
    headers = [
        'Ser.No', 'TIN', 'Name of Employee', 'Position Held', 'Residency',
        'Basic Salary', 'Secondary Employment', 'Paid SSNIT', 'Social Security Fund',
        'Third Tier', 'Cash Allowance', 'Bonus Income', 'Final Tax on Bonus',
        'Excess Bonus', 'Total Cash Emolument', 'Accommodation Element',
        'Vehicle Element', 'Non-Cash Benefit', 'Total Assessable Income',
        'Deductible Reliefs', 'Total Reliefs', 'Chargeable Income',
        'Tax Deductible', 'Overtime Income', 'Overtime Tax',
        'Total Tax Payable to GRA', 'Severance Pay', 'Remarks'
    ]

    header_row = 8
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows start at row 9
    data_start_row = 9

    # Write data
    for row_idx, row_data in enumerate(data, data_start_row):
        values = [
            row_data.get('ser_no', ''),
            row_data.get('tin_number', ''),
            row_data.get('employee_name', ''),
            row_data.get('position', ''),
            row_data.get('residency', ''),
            row_data.get('basic_salary', 0),
            row_data.get('secondary_employment', 'N'),
            row_data.get('paid_ssnit', 'Y'),
            row_data.get('social_security_fund', 0),
            row_data.get('third_tier', 0),
            row_data.get('cash_allowance', 0),
            row_data.get('bonus_income', 0),
            row_data.get('final_tax_on_bonus', 0),
            row_data.get('excess_bonus', 0),
            row_data.get('total_cash_emolument', 0),
            row_data.get('accommodation_element', 0),
            row_data.get('vehicle_element', 0),
            row_data.get('non_cash_benefit', 0),
            row_data.get('total_assessable_income', 0),
            row_data.get('deductible_reliefs', 0),
            row_data.get('total_reliefs', 0),
            row_data.get('chargeable_income', 0),
            row_data.get('tax_deductible', 0),
            row_data.get('overtime_income', 0),
            row_data.get('overtime_tax', 0),
            row_data.get('total_tax_payable', 0),
            row_data.get('severance_pay', 0),
            row_data.get('remarks', ''),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # Format numeric columns
            if col_idx in [6, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27]:
                cell.alignment = right_align
                cell.number_format = '#,##0.00'
            elif col_idx in [1]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Totals row
    totals_row = data_start_row + len(data)
    ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True, size=9)
    ws.cell(row=totals_row, column=1).fill = total_fill
    ws.cell(row=totals_row, column=1).border = thin_border

    # Calculate and write totals for numeric columns
    numeric_cols = {
        6: 'basic_salary', 9: 'social_security_fund', 10: 'third_tier',
        11: 'cash_allowance', 12: 'bonus_income', 13: 'final_tax_on_bonus',
        14: 'excess_bonus', 15: 'total_cash_emolument', 16: 'accommodation_element',
        17: 'vehicle_element', 18: 'non_cash_benefit', 19: 'total_assessable_income',
        20: 'deductible_reliefs', 21: 'total_reliefs', 22: 'chargeable_income',
        23: 'tax_deductible', 24: 'overtime_income', 25: 'overtime_tax',
        26: 'total_tax_payable', 27: 'severance_pay'
    }

    for col_idx, field in numeric_cols.items():
        total = sum(float(row.get(field, 0) or 0) for row in data)
        cell = ws.cell(row=totals_row, column=col_idx, value=total)
        cell.font = Font(bold=True, size=9)
        cell.fill = total_fill
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = right_align

    # Fill remaining total row cells with border
    for col_idx in range(2, len(headers) + 1):
        if col_idx not in numeric_cols:
            cell = ws.cell(row=totals_row, column=col_idx)
            cell.fill = total_fill
            cell.border = thin_border

    # Adjust column widths
    column_widths = [6, 12, 25, 18, 10, 12, 8, 8, 12, 10, 12, 12, 12, 10, 14,
                     14, 12, 12, 14, 12, 12, 14, 12, 12, 10, 14, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_paye_gra_pdf(data: list, org_name: str, period: str, filename: str) -> HttpResponse:
    """
    Generate GRA-compliant PAYE PDF report.
    Uses landscape A3 for wider format to fit all columns.
    Uses Paragraph objects for proper text wrapping to prevent overflow.
    """
    from reportlab.lib.pagesizes import A3

    buffer = io.BytesIO()

    # Use landscape A3 for more space
    page_size = landscape(A3)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.25*inch,
        leftMargin=0.25*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title section
    title_style = ParagraphStyle(
        'GRATitle', parent=styles['Heading1'], fontSize=14,
        alignment=TA_CENTER, spaceAfter=5
    )
    subtitle_style = ParagraphStyle(
        'GRASubtitle', parent=styles['Normal'], fontSize=11,
        alignment=TA_CENTER, spaceAfter=5
    )
    org_style = ParagraphStyle(
        'OrgName', parent=styles['Heading2'], fontSize=12,
        alignment=TA_CENTER, spaceAfter=5, fontName='Helvetica-Bold'
    )

    # Cell styles for text wrapping
    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontSize=6,
        leading=7, wordWrap='CJK'
    )
    cell_style_right = ParagraphStyle(
        'CellRight', parent=styles['Normal'], fontSize=6,
        leading=7, wordWrap='CJK', alignment=TA_RIGHT
    )
    cell_style_center = ParagraphStyle(
        'CellCenter', parent=styles['Normal'], fontSize=6,
        leading=7, wordWrap='CJK', alignment=TA_CENTER
    )
    cell_style_bold = ParagraphStyle(
        'CellBold', parent=styles['Normal'], fontSize=6,
        leading=7, fontName='Helvetica-Bold', wordWrap='CJK'
    )
    header_style = ParagraphStyle(
        'HeaderCell', parent=styles['Normal'], fontSize=6,
        leading=7, fontName='Helvetica-Bold', alignment=TA_CENTER,
        textColor=colors.whitesmoke, wordWrap='CJK'
    )

    elements.append(Paragraph("GHANA REVENUE AUTHORITY", title_style))
    elements.append(Paragraph("DOMESTIC TAX REVENUE DIVISION", subtitle_style))
    elements.append(Paragraph("PAYE MONTHLY TAX DEDUCTIONS SCHEDULE", title_style))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph(org_name, org_style))
    elements.append(Paragraph(f"Period: {period}" if period else "Period: N/A", subtitle_style))
    elements.append(Spacer(1, 0.15*inch))

    # Column headers (abbreviated for PDF space) - use Paragraph objects
    header_texts = [
        'No', 'TIN', 'Name', 'Position', 'Res', 'Basic', 'Sec<br/>Emp', 'SSNIT',
        'SSF', 'T3', 'Allow', 'Bonus', 'Bonus<br/>Tax', 'Exc<br/>Bonus', 'Cash<br/>Emol',
        'Accom', 'Veh', 'Non<br/>Cash', 'Assess<br/>Inc', 'Reliefs', 'Tot<br/>Rel',
        'Charge<br/>Inc', 'Tax<br/>Ded', 'OT Inc', 'OT Tax', 'Tot Tax', 'Sev', 'Rmk'
    ]
    headers = [Paragraph(h, header_style) for h in header_texts]

    # Prepare table data with Paragraph objects
    table_data = [headers]

    for row in data:
        # Text columns use cell_style, numeric use cell_style_right
        row_values = [
            Paragraph(str(row.get('ser_no', '')), cell_style_center),
            Paragraph(str(row.get('tin_number', '') or ''), cell_style),
            Paragraph(str(row.get('employee_name', '') or ''), cell_style),
            Paragraph(str(row.get('position', '') or ''), cell_style),
            Paragraph(str(row.get('residency', '') or '')[:3], cell_style_center),
            Paragraph(f"{float(row.get('basic_salary', 0)):,.0f}", cell_style_right),
            Paragraph(str(row.get('secondary_employment', 'N')), cell_style_center),
            Paragraph(str(row.get('paid_ssnit', 'Y')), cell_style_center),
            Paragraph(f"{float(row.get('social_security_fund', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('third_tier', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('cash_allowance', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('bonus_income', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('final_tax_on_bonus', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('excess_bonus', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('total_cash_emolument', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('accommodation_element', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('vehicle_element', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('non_cash_benefit', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('total_assessable_income', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('deductible_reliefs', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('total_reliefs', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('chargeable_income', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('tax_deductible', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('overtime_income', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('overtime_tax', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('total_tax_payable', 0)):,.0f}", cell_style_right),
            Paragraph(f"{float(row.get('severance_pay', 0)):,.0f}", cell_style_right),
            Paragraph(str(row.get('remarks', '') or ''), cell_style),
        ]
        table_data.append(row_values)

    # Add totals row with Paragraph objects
    totals = [
        Paragraph('TOTALS', cell_style_bold),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
        Paragraph('', cell_style),
    ]

    # Calculate totals for numeric columns
    basic_total = sum(float(row.get('basic_salary', 0) or 0) for row in data)
    totals.append(Paragraph(f"{basic_total:,.0f}", cell_style_bold))
    totals.append(Paragraph('', cell_style))  # sec_emp
    totals.append(Paragraph('', cell_style))  # paid_ssnit

    numeric_fields = [
        'social_security_fund', 'third_tier', 'cash_allowance', 'bonus_income',
        'final_tax_on_bonus', 'excess_bonus', 'total_cash_emolument',
        'accommodation_element', 'vehicle_element', 'non_cash_benefit',
        'total_assessable_income', 'deductible_reliefs', 'total_reliefs',
        'chargeable_income', 'tax_deductible', 'overtime_income', 'overtime_tax',
        'total_tax_payable', 'severance_pay'
    ]

    for field in numeric_fields:
        total = sum(float(row.get(field, 0) or 0) for row in data)
        totals.append(Paragraph(f"{total:,.0f}", cell_style_bold))

    totals.append(Paragraph('', cell_style))  # remarks
    table_data.append(totals)

    # Column widths
    col_widths = [
        0.28*inch,   # No
        0.55*inch,   # TIN
        0.95*inch,   # Name
        0.75*inch,   # Position
        0.28*inch,   # Res
        0.52*inch,   # Basic
        0.28*inch,   # Sec Emp
        0.32*inch,   # SSNIT
        0.45*inch,   # SSF
        0.35*inch,   # T3
        0.45*inch,   # Allow
        0.42*inch,   # Bonus
        0.42*inch,   # Bonus Tax
        0.42*inch,   # Exc Bonus
        0.52*inch,   # Cash Emol
        0.42*inch,   # Accom
        0.35*inch,   # Veh
        0.42*inch,   # Non Cash
        0.52*inch,   # Assess Inc
        0.45*inch,   # Reliefs
        0.42*inch,   # Tot Rel
        0.52*inch,   # Charge Inc
        0.45*inch,   # Tax Ded
        0.42*inch,   # OT Inc
        0.42*inch,   # OT Tax
        0.52*inch,   # Tot Tax
        0.35*inch,   # Sev
        0.42*inch,   # Rmk
    ]

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style - simpler since Paragraph handles alignment
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        # Totals row styling
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#BDD7EE')),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_paye_gra_report(payroll_run_id: str = None, filters: dict = None, format: str = 'excel') -> HttpResponse:
    """Export GRA-compliant PAYE report in the specified format."""
    data, org_name, period = get_paye_gra_data(payroll_run_id, filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not data:
        # Return empty response with message
        if format == 'excel':
            return generate_excel_response([], ['No Data Available'], f'paye_gra_{timestamp}.xlsx', 'GRA PAYE Report')
        elif format == 'pdf':
            return generate_pdf_response([], ['No Data Available'], f'paye_gra_{timestamp}.pdf', 'GRA PAYE Report')
        else:
            return generate_csv_response([], ['No Data Available'], f'paye_gra_{timestamp}.csv')

    if format == 'excel':
        return export_paye_gra_excel(data, org_name or 'ORGANIZATION', period or '', f'paye_gra_{timestamp}.xlsx')
    elif format == 'pdf':
        return export_paye_gra_pdf(data, org_name or 'ORGANIZATION', period or '', f'paye_gra_{timestamp}.pdf')
    else:
        # CSV format - use standard headers
        headers = [
            'Ser.No', 'TIN', 'Name of Employee', 'Position Held', 'Residency',
            'Basic Salary', 'Secondary Employment', 'Paid SSNIT', 'Social Security Fund',
            'Third Tier', 'Cash Allowance', 'Bonus Income', 'Final Tax on Bonus',
            'Excess Bonus', 'Total Cash Emolument', 'Accommodation Element',
            'Vehicle Element', 'Non-Cash Benefit', 'Total Assessable Income',
            'Deductible Reliefs', 'Total Reliefs', 'Chargeable Income',
            'Tax Deductible', 'Overtime Income', 'Overtime Tax',
            'Total Tax Payable to GRA', 'Severance Pay', 'Remarks'
        ]
        return generate_csv_response(data, headers, f'paye_gra_{timestamp}.csv')


def get_ssnit_data(payroll_run_id: str = None, filters: dict = None):
    """Get SSNIT contribution report data."""
    if payroll_run_id:
        items = PayrollItem.objects.filter(payroll_run_id=payroll_run_id)
    else:
        latest_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()
        if latest_run:
            items = PayrollItem.objects.filter(payroll_run=latest_run)
        else:
            items = PayrollItem.objects.none()

    items = items.select_related(
        'employee', 'employee__department', 'employee__division',
        'employee__directorate', 'employee__position', 'employee__grade'
    )

    # Apply filters
    items = apply_payroll_item_filters(items, filters)

    headers = [
        'Employee Number', 'Employee Name', 'SSNIT Number',
        'Basic Salary', 'SSNIT Employee 5.5', 'SSNIT Employer 13',
        'Tier 2 Employer 5', 'Total Contribution'
    ]

    data = []
    for item in items:
        total = float(item.ssnit_employee) + float(item.ssnit_employer) + float(item.tier2_employer)
        data.append({
            'employee_number': item.employee.employee_number,
            'employee_name': item.employee.full_name,
            'ssnit_number': item.employee.ssnit_number or '',
            'basic_salary': float(item.basic_salary),
            'ssnit_employee_5.5': float(item.ssnit_employee),
            'ssnit_employer_13': float(item.ssnit_employer),
            'tier_2_employer_5': float(item.tier2_employer),
            'total_contribution': total,
        })

    return data, headers


def get_bank_advice_data(payroll_run_id: str = None, filters: dict = None):
    """Get bank advice report data."""
    if payroll_run_id:
        items = PayrollItem.objects.filter(payroll_run_id=payroll_run_id)
    else:
        latest_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()
        if latest_run:
            items = PayrollItem.objects.filter(payroll_run=latest_run)
        else:
            items = PayrollItem.objects.none()

    items = items.select_related(
        'employee', 'payroll_run', 'employee__department', 'employee__division',
        'employee__directorate', 'employee__position', 'employee__grade'
    ).filter(
        bank_account_number__isnull=False
    ).order_by('bank_name', 'employee__employee_number')

    # Apply filters
    items = apply_payroll_item_filters(items, filters)

    headers = [
        'Bank', 'Employee Number', 'Employee Name',
        'Account Number', 'Branch', 'Net Salary', 'Reference'
    ]

    data = []
    for item in items:
        data.append({
            'bank': item.bank_name or '',
            'employee_number': item.employee.employee_number,
            'employee_name': item.employee.full_name,
            'account_number': item.bank_account_number or '',
            'branch': item.bank_branch or '',
            'net_salary': float(item.net_salary),
            'reference': item.payment_reference or f'{item.payroll_run.run_number}-{item.employee.employee_number}',
        })

    return data, headers


def get_leave_balance_data(filters: dict = None):
    """Get leave balance report data."""
    from django.utils import timezone
    year = filters.get('year', timezone.now().year) if filters else timezone.now().year

    balances = LeaveBalance.objects.filter(
        year=year
    ).select_related(
        'employee', 'leave_type', 'employee__department',
        'employee__division', 'employee__directorate',
        'employee__position', 'employee__grade', 'employee__staff_category'
    )

    # Apply employee filters via employee relation
    if filters:
        if filters.get('employee_code'):
            balances = balances.filter(employee__employee_number__icontains=filters['employee_code'])
        if filters.get('division'):
            balances = balances.filter(employee__division_id=filters['division'])
        if filters.get('directorate'):
            balances = balances.filter(employee__directorate_id=filters['directorate'])
        if filters.get('department'):
            balances = balances.filter(employee__department_id=filters['department'])
        if filters.get('position'):
            balances = balances.filter(employee__position_id=filters['position'])
        if filters.get('grade'):
            balances = balances.filter(employee__grade_id=filters['grade'])
        if filters.get('staff_category'):
            balances = balances.filter(employee__staff_category_id=filters['staff_category'])

    headers = [
        'Employee Number', 'Employee Name', 'Department', 'Leave Type',
        'Opening Balance', 'Earned', 'Taken', 'Pending', 'Available'
    ]

    data = []
    for bal in balances:
        available = (bal.opening_balance or 0) + (bal.earned or 0) - (bal.taken or 0) - (bal.pending or 0)
        data.append({
            'employee_number': bal.employee.employee_number,
            'employee_name': bal.employee.full_name,
            'department': bal.employee.department.name if bal.employee.department else '',
            'leave_type': bal.leave_type.name if bal.leave_type else '',
            'opening_balance': float(bal.opening_balance or 0),
            'earned': float(bal.earned or 0),
            'taken': float(bal.taken or 0),
            'pending': float(bal.pending or 0),
            'available': float(available),
        })

    return data, headers


def get_loan_outstanding_data(filters: dict = None):
    """Get outstanding loans report data."""
    queryset = LoanAccount.objects.filter(
        status__in=['ACTIVE', 'DISBURSED']
    ).select_related(
        'employee', 'loan_type', 'employee__department',
        'employee__division', 'employee__directorate',
        'employee__position', 'employee__grade', 'employee__staff_category'
    )

    # Apply employee filters via employee relation
    if filters:
        if filters.get('employee_code'):
            queryset = queryset.filter(employee__employee_number__icontains=filters['employee_code'])
        if filters.get('division'):
            queryset = queryset.filter(employee__division_id=filters['division'])
        if filters.get('directorate'):
            queryset = queryset.filter(employee__directorate_id=filters['directorate'])
        if filters.get('department'):
            queryset = queryset.filter(employee__department_id=filters['department'])
        if filters.get('position'):
            queryset = queryset.filter(employee__position_id=filters['position'])
        if filters.get('grade'):
            queryset = queryset.filter(employee__grade_id=filters['grade'])
        if filters.get('staff_category'):
            queryset = queryset.filter(employee__staff_category_id=filters['staff_category'])

    headers = [
        'Loan Number', 'Employee Number', 'Employee Name', 'Department',
        'Loan Type', 'Principal Amount', 'Disbursed Amount',
        'Principal Paid', 'Outstanding Balance', 'Monthly Installment'
    ]

    data = []
    for loan in queryset:
        data.append({
            'loan_number': loan.loan_number,
            'employee_number': loan.employee.employee_number,
            'employee_name': loan.employee.full_name,
            'department': loan.employee.department.name if loan.employee.department else '',
            'loan_type': loan.loan_type.name if loan.loan_type else '',
            'principal_amount': float(loan.principal_amount or 0),
            'disbursed_amount': float(loan.disbursed_amount or 0),
            'principal_paid': float(loan.principal_paid or 0),
            'outstanding_balance': float(loan.outstanding_balance or 0),
            'monthly_installment': float(loan.monthly_installment or 0),
        })

    return data, headers


# =============================================================================
# Export Functions (CSV, Excel, PDF)
# =============================================================================

def export_employee_master(filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export employee master report."""
    data, headers = get_employee_master_data(filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Employee Master Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'employee_master_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'employee_master_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'employee_master_{timestamp}.csv')


def export_headcount(filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export headcount report with chart visualizations for PDF/Excel."""
    data, headers = get_headcount_data(filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Headcount Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'headcount_{timestamp}.csv')

    # Gather chart breakdown data
    queryset = Employee.objects.filter(status='ACTIVE')
    queryset = apply_employee_filters(queryset, filters)

    by_division = [
        {'name': row['name'] or 'Unassigned', 'value': row['value']}
        for row in queryset.values(name=F('division__name')).annotate(value=Count('id')).order_by('-value')
    ]
    by_directorate = [
        {'name': row['name'] or 'Unassigned', 'value': row['value']}
        for row in queryset.values(name=F('directorate__name')).annotate(value=Count('id')).order_by('-value')
    ]
    by_grade = [
        {'name': row['name'] or 'Unassigned', 'value': row['value']}
        for row in queryset.values(name=F('grade__name')).annotate(value=Count('id')).order_by('-value')
    ]
    by_location = [
        {'name': row['name'] or 'Unassigned', 'value': row['value']}
        for row in queryset.values(name=F('work_location__name')).annotate(value=Count('id')).order_by('-value')
    ]
    by_emp_type = [
        {'name': row['name'] or 'Not Specified', 'value': row['value']}
        for row in queryset.values(name=F('employment_type')).annotate(value=Count('id')).order_by('-value')
    ]

    total = str(queryset.count())
    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'Headcount by Department'},
        {'type': 'chart', 'chart_type': 'pie', 'data': by_division, 'title': 'Headcount by Division',
         'options': {'donut': True, 'center_label': total}},
        {'type': 'chart', 'chart_type': 'pie', 'data': by_directorate, 'title': 'Headcount by Directorate',
         'options': {'donut': True, 'center_label': total}},
        {'type': 'chart', 'chart_type': 'pie', 'data': by_grade, 'title': 'Headcount by Grade',
         'options': {'donut': True, 'center_label': total}},
        {'type': 'chart', 'chart_type': 'pie', 'data': by_location, 'title': 'Headcount by Location',
         'options': {'donut': True, 'center_label': total}},
        {'type': 'chart', 'chart_type': 'bar', 'data': by_emp_type, 'title': 'Headcount by Employment Type',
         'options': {}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'headcount_{timestamp}.pdf', title)
    else:
        return generate_excel_with_charts_response(sections, f'headcount_{timestamp}.xlsx', title)


def export_payroll_summary(payroll_run_id: str = None, filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export payroll summary."""
    data, headers = get_payroll_summary_data(payroll_run_id, filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Payroll Summary Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'payroll_summary_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'payroll_summary_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'payroll_summary_{timestamp}.csv')


def export_paye_report(payroll_run_id: str = None, filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export PAYE tax report."""
    data, headers = get_paye_data(payroll_run_id, filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "PAYE Tax Report (GRA)"

    if format == 'excel':
        return generate_excel_response(data, headers, f'paye_report_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'paye_report_{timestamp}.pdf', title)
    else:
        return generate_csv_response(data, headers, f'paye_report_{timestamp}.csv')


def export_ssnit_report(payroll_run_id: str = None, filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export SSNIT contribution report."""
    data, headers = get_ssnit_data(payroll_run_id, filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "SSNIT Contribution Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'ssnit_report_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'ssnit_report_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'ssnit_report_{timestamp}.csv')


def export_bank_advice(payroll_run_id: str = None, filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export bank advice report."""
    data, headers = get_bank_advice_data(payroll_run_id, filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Bank Advice Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'bank_advice_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'bank_advice_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'bank_advice_{timestamp}.csv')


def export_leave_balance(filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export leave balance report with chart visualizations for PDF/Excel."""
    data, headers = get_leave_balance_data(filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Leave Balance Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'leave_balance_{timestamp}.csv')

    # Aggregate leave taken by type for bar chart
    by_leave_type = {}
    for row in data:
        lt = row.get('leave_type', 'Unknown')
        by_leave_type[lt] = by_leave_type.get(lt, 0) + float(row.get('taken', 0))

    leave_type_chart = [{'name': k, 'value': v} for k, v in
                        sorted(by_leave_type.items(), key=lambda x: -x[1])]

    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'Leave Balances'},
        {'type': 'chart', 'chart_type': 'bar', 'data': leave_type_chart,
         'title': 'Leave Days Taken by Type', 'options': {}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'leave_balance_{timestamp}.pdf', title,
                                                  landscape_mode=True)
    else:
        return generate_excel_with_charts_response(sections, f'leave_balance_{timestamp}.xlsx', title)


def export_loan_outstanding(filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export outstanding loans report."""
    data, headers = get_loan_outstanding_data(filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Outstanding Loans Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'outstanding_loans_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'outstanding_loans_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'outstanding_loans_{timestamp}.csv')


def get_payroll_master_data(payroll_run_id: str = None, filters: dict = None):
    """
    Get Payroll Master Report data with detailed breakdown per employee.
    Shows ALL individual transactions for earnings, deductions, and employer contributions.
    """
    from collections import OrderedDict

    if payroll_run_id:
        try:
            payroll_run = PayrollRun.objects.get(id=payroll_run_id)
        except PayrollRun.DoesNotExist:
            payroll_run = None
    else:
        payroll_run = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date').first()

    if not payroll_run:
        return [], ['No Data']

    items = PayrollItem.objects.filter(
        payroll_run=payroll_run
    ).select_related(
        'employee', 'employee__department', 'employee__position',
        'employee__division', 'employee__directorate', 'employee__grade'
    ).prefetch_related(
        'details', 'details__pay_component'
    ).order_by('employee__employee_number')

    # Apply filters
    items = apply_payroll_item_filters(items, filters)
    items_list = list(items)

    if not items_list:
        return [], ['No Data']

    # First pass: collect all unique pay components by type (ordered by display_order)
    earning_components = OrderedDict()  # code -> name
    deduction_components = OrderedDict()
    employer_components = OrderedDict()

    for item in items_list:
        for detail in item.details.all():
            comp = detail.pay_component
            if comp.component_type == 'EARNING':
                if comp.code not in earning_components:
                    earning_components[comp.code] = comp.name
            elif comp.component_type == 'DEDUCTION':
                if comp.code not in deduction_components:
                    deduction_components[comp.code] = comp.name
            elif comp.component_type == 'EMPLOYER':
                if comp.code not in employer_components:
                    employer_components[comp.code] = comp.name

    # Add standard SSNIT/PAYE if not already in details (from PayrollItem fields)
    if 'SSNIT_EMP' not in deduction_components:
        deduction_components['SSNIT_EMP'] = 'SSNIT Employee (5.5%)'
    if 'PAYE' not in deduction_components:
        deduction_components['PAYE'] = 'PAYE Tax'
    if 'SSNIT_EMPLOYER' not in employer_components:
        employer_components['SSNIT_EMPLOYER'] = 'SSNIT Employer (13%)'
    if 'TIER2_EMPLOYER' not in employer_components:
        employer_components['TIER2_EMPLOYER'] = 'Tier 2 Employer (5%)'

    # Build headers list and corresponding keys for data extraction
    headers = []
    column_keys = []  # Parallel list to track how to get data for each column

    # Basic employee info
    headers.extend(['Employee ID', 'Employee Name', 'Department', 'Position'])
    column_keys.extend(['employee_id', 'employee_name', 'department', 'position'])

    # Earnings section - each earning component as its own column
    earning_codes = list(earning_components.keys())
    for code in earning_codes:
        headers.append(earning_components[code])
        column_keys.append(f'earning_{code}')
    headers.append('GROSS SALARY')
    column_keys.append('gross_salary')

    # Deductions section - each deduction component as its own column
    deduction_codes = list(deduction_components.keys())
    for code in deduction_codes:
        headers.append(deduction_components[code])
        column_keys.append(f'deduction_{code}')
    headers.append('TOTAL DEDUCTIONS')
    column_keys.append('total_deductions')

    # Net salary
    headers.append('NET SALARY')
    column_keys.append('net_salary')

    # Employer contributions section - each employer component as its own column
    employer_codes = list(employer_components.keys())
    for code in employer_codes:
        headers.append(employer_components[code])
        column_keys.append(f'employer_{code}')
    headers.append('TOTAL EMPLOYER COST')
    column_keys.append('total_employer_cost')

    # Second pass: build data rows
    data = []
    for item in items_list:
        # Create a dict to hold component amounts by code
        component_amounts = {}
        for detail in item.details.all():
            component_amounts[detail.pay_component.code] = float(detail.amount)

        # Build row using column_keys for consistent mapping
        row = {}

        # Basic info
        row['employee_id'] = item.employee.employee_number
        row['employee_name'] = item.employee.full_name
        row['department'] = item.employee.department.name if item.employee.department else ''
        row['position'] = item.employee.position.title if item.employee.position else ''

        # Earnings
        for code in earning_codes:
            row[f'earning_{code}'] = component_amounts.get(code, 0.0)
        row['gross_salary'] = float(item.gross_earnings)

        # Deductions
        for code in deduction_codes:
            if code == 'SSNIT_EMP':
                # Use PayrollItem field value, fall back to detail
                row[f'deduction_{code}'] = float(item.ssnit_employee) if item.ssnit_employee else component_amounts.get(code, 0.0)
            elif code == 'PAYE':
                row[f'deduction_{code}'] = float(item.paye) if item.paye else component_amounts.get(code, 0.0)
            else:
                row[f'deduction_{code}'] = component_amounts.get(code, 0.0)
        row['total_deductions'] = float(item.total_deductions)

        # Net salary
        row['net_salary'] = float(item.net_salary)

        # Employer contributions
        for code in employer_codes:
            if code == 'SSNIT_EMPLOYER':
                row[f'employer_{code}'] = float(item.ssnit_employer) if item.ssnit_employer else component_amounts.get(code, 0.0)
            elif code == 'TIER2_EMPLOYER':
                row[f'employer_{code}'] = float(item.tier2_employer) if item.tier2_employer else component_amounts.get(code, 0.0)
            else:
                row[f'employer_{code}'] = component_amounts.get(code, 0.0)
        row['total_employer_cost'] = float(item.employer_cost)

        data.append(row)

    return data, headers, column_keys


def export_payroll_master(payroll_run_id: str = None, filters: dict = None, format: str = 'csv') -> HttpResponse:
    """Export Payroll Master Report with all individual transactions."""
    result = get_payroll_master_data(payroll_run_id, filters)

    # Handle the case where no data is found
    if len(result) == 2:
        data, headers = result
        column_keys = None
    else:
        data, headers, column_keys = result

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Payroll Master Report"

    if not data or headers == ['No Data']:
        # Return empty response
        if format == 'excel':
            return generate_excel_response([], ['No Data Available'], f'payroll_master_{timestamp}.xlsx', title)
        elif format == 'pdf':
            return generate_pdf_response([], ['No Data Available'], f'payroll_master_{timestamp}.pdf', title)
        else:
            return generate_csv_response([], ['No Data Available'], f'payroll_master_{timestamp}.csv')

    # Use custom generators that use column_keys for proper data mapping
    if format == 'excel':
        return generate_payroll_master_excel(data, headers, column_keys, f'payroll_master_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_payroll_master_pdf(data, headers, column_keys, f'payroll_master_{timestamp}.pdf', title)
    else:
        return generate_payroll_master_csv(data, headers, column_keys, f'payroll_master_{timestamp}.csv')


def generate_payroll_master_csv(data: list, headers: list, column_keys: list, filename: str) -> HttpResponse:
    """Generate CSV for Payroll Master Report using column_keys mapping."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(headers)

    for row in data:
        row_data = []
        for key in column_keys:
            value = row.get(key, '')
            if isinstance(value, float):
                row_data.append(value)
            else:
                row_data.append(value)
        writer.writerow(row_data)

    return response


def generate_payroll_master_excel(data: list, headers: list, column_keys: list, filename: str, title: str = None) -> HttpResponse:
    """Generate Excel for Payroll Master Report using column_keys mapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Master"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    earning_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    deduction_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    employer_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 1

    # Add title if provided
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        start_row = 3

    # Write headers with color coding
    for col, (header, key) in enumerate(zip(headers, column_keys), 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        # Color code by section
        if key.startswith('earning_') or header == 'GROSS SALARY':
            cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Dark green
        elif key.startswith('deduction_') or header == 'TOTAL DEDUCTIONS':
            cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red
        elif key.startswith('employer_') or header == 'TOTAL EMPLOYER COST':
            cell.fill = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")  # Dark goldenrod
        elif header == 'NET SALARY':
            cell.fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # Indigo
        else:
            cell.fill = header_fill

    # Write data
    for row_idx, row in enumerate(data, start_row + 1):
        for col_idx, key in enumerate(column_keys, 1):
            value = row.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'

                # Light background colors for data cells
                if key.startswith('earning_') or key == 'gross_salary':
                    cell.fill = earning_fill
                elif key.startswith('deduction_') or key == 'total_deductions':
                    cell.fill = deduction_fill
                elif key.startswith('employer_') or key == 'total_employer_cost':
                    cell.fill = employer_fill
                elif key == 'net_salary':
                    cell.fill = total_fill

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(start_row, len(data) + start_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def get_payroll_reconciliation_data(current_run_id: str = None, previous_run_id: str = None):
    """
    Get payroll reconciliation data between two payroll runs.
    Returns data suitable for export in CSV/Excel/PDF format.
    """
    from django.utils import timezone

    # Get runs - either from params or last two approved runs
    if current_run_id and previous_run_id:
        try:
            current_run = PayrollRun.objects.get(id=current_run_id)
            previous_run = PayrollRun.objects.get(id=previous_run_id)
        except PayrollRun.DoesNotExist:
            return [], [], None
    else:
        runs = PayrollRun.objects.filter(
            status__in=['COMPUTED', 'APPROVED', 'PAID']
        ).order_by('-run_date')[:2]

        if len(runs) < 2:
            return [], [], None

        current_run = runs[0]
        previous_run = runs[1]

    # Get items from both runs
    current_items = {
        item.employee_id: item
        for item in PayrollItem.objects.filter(payroll_run=current_run).select_related(
            'employee', 'employee__department', 'employee__position', 'employee__grade'
        )
    }
    previous_items = {
        item.employee_id: item
        for item in PayrollItem.objects.filter(payroll_run=previous_run).select_related(
            'employee', 'employee__department', 'employee__position', 'employee__grade'
        )
    }

    # All employee IDs from both periods
    all_employee_ids = set(current_items.keys()) | set(previous_items.keys())

    def safe_float(val):
        if val is None:
            return 0.0
        return float(val)

    # Build reconciliation data
    data = []
    for emp_id in all_employee_ids:
        current = current_items.get(emp_id)
        previous = previous_items.get(emp_id)

        if current and not previous:
            # New employee
            emp = current.employee
            data.append({
                'status': 'NEW',
                'employee_number': emp.employee_number,
                'employee_name': emp.full_name,
                'department': emp.department.name if emp.department else '',
                'prev_gross': 0,
                'curr_gross': safe_float(current.gross_earnings),
                'gross_diff': safe_float(current.gross_earnings),
                'prev_net': 0,
                'curr_net': safe_float(current.net_salary),
                'net_diff': safe_float(current.net_salary),
                'reason': 'New Hire',
            })
        elif previous and not current:
            # Separated employee
            emp = previous.employee
            data.append({
                'status': 'SEPARATED',
                'employee_number': emp.employee_number,
                'employee_name': emp.full_name,
                'department': emp.department.name if emp.department else '',
                'prev_gross': safe_float(previous.gross_earnings),
                'curr_gross': 0,
                'gross_diff': -safe_float(previous.gross_earnings),
                'prev_net': safe_float(previous.net_salary),
                'curr_net': 0,
                'net_diff': -safe_float(previous.net_salary),
                'reason': emp.exit_reason or 'Separation',
            })
        else:
            # Compare current and previous
            emp = current.employee
            curr_gross = safe_float(current.gross_earnings)
            prev_gross = safe_float(previous.gross_earnings)
            curr_net = safe_float(current.net_salary)
            prev_net = safe_float(previous.net_salary)

            gross_diff = curr_gross - prev_gross
            net_diff = curr_net - prev_net

            # Check if there's any meaningful change (>0.01)
            has_change = abs(gross_diff) > 0.01 or abs(net_diff) > 0.01

            if has_change:
                # Determine change reason
                curr_basic = safe_float(current.basic_salary)
                prev_basic = safe_float(previous.basic_salary)
                basic_diff = curr_basic - prev_basic

                reasons = []
                if abs(basic_diff) > 0.01:
                    reasons.append('Salary Change')
                if abs(gross_diff - basic_diff) > 0.01:
                    reasons.append('Allowances')
                if abs(net_diff - gross_diff) > 0.01:
                    reasons.append('Deductions')

                data.append({
                    'status': 'CHANGED',
                    'employee_number': emp.employee_number,
                    'employee_name': emp.full_name,
                    'department': emp.department.name if emp.department else '',
                    'prev_gross': prev_gross,
                    'curr_gross': curr_gross,
                    'gross_diff': gross_diff,
                    'prev_net': prev_net,
                    'curr_net': curr_net,
                    'net_diff': net_diff,
                    'reason': ', '.join(reasons) if reasons else 'Adjustment',
                })
            else:
                data.append({
                    'status': 'UNCHANGED',
                    'employee_number': emp.employee_number,
                    'employee_name': emp.full_name,
                    'department': emp.department.name if emp.department else '',
                    'prev_gross': prev_gross,
                    'curr_gross': curr_gross,
                    'gross_diff': gross_diff,
                    'prev_net': prev_net,
                    'curr_net': curr_net,
                    'net_diff': net_diff,
                    'reason': 'No Change',
                })

    # Sort by status priority: NEW, SEPARATED, CHANGED, UNCHANGED
    status_order = {'NEW': 0, 'SEPARATED': 1, 'CHANGED': 2, 'UNCHANGED': 3}
    data.sort(key=lambda x: (status_order.get(x['status'], 99), x['employee_name']))

    headers = [
        'Status', 'Employee Number', 'Employee Name', 'Department',
        'Prev Gross', 'Curr Gross', 'Gross Diff',
        'Prev Net', 'Curr Net', 'Net Diff', 'Reason'
    ]

    current_period = current_run.payroll_period.name if current_run.payroll_period else current_run.run_number
    previous_period = previous_run.payroll_period.name if previous_run.payroll_period else previous_run.run_number
    title = f"Payroll Reconciliation: {previous_period} vs {current_period}"

    return data, headers, title


def generate_payroll_master_pdf(data: list, headers: list, column_keys: list, filename: str, title: str = None) -> HttpResponse:
    """Generate PDF for Payroll Master Report using column_keys mapping with text wrapping."""
    buffer = io.BytesIO()

    # Always use landscape for payroll master (many columns)
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.3*inch,
        leftMargin=0.3*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )

    elements = []
    styles = getSampleStyleSheet()

    # Title
    if title:
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=TA_CENTER,
            spaceAfter=10
        )
        elements.append(Paragraph(title, title_style))

    # Subtitle with date
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        spaceAfter=10
    )
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.1*inch))

    # Cell styles for text wrapping - different colors for different sections
    def make_header_style(bg_color):
        return ParagraphStyle(
            f'Header_{bg_color}', parent=styles['Normal'], fontSize=6,
            leading=7, fontName='Helvetica-Bold', alignment=TA_CENTER,
            textColor=colors.whitesmoke, wordWrap='CJK'
        )

    cell_style = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontSize=6,
        leading=7, wordWrap='CJK'
    )
    cell_style_right = ParagraphStyle(
        'CellRight', parent=styles['Normal'], fontSize=6,
        leading=7, wordWrap='CJK', alignment=TA_RIGHT
    )

    # Create header row with Paragraph objects
    header_row = []
    header_bg_colors = []
    for idx, (header, key) in enumerate(zip(headers, column_keys)):
        # Determine background color based on section
        if key.startswith('earning_') or key == 'gross_salary':
            bg_color = '#006400'  # Dark green
        elif key.startswith('deduction_') or key == 'total_deductions':
            bg_color = '#8B0000'  # Dark red
        elif key.startswith('employer_') or key == 'total_employer_cost':
            bg_color = '#B8860B'  # Dark goldenrod
        elif key == 'net_salary':
            bg_color = '#4B0082'  # Indigo
        else:
            bg_color = '#4472C4'  # Blue

        header_bg_colors.append(bg_color)
        header_row.append(Paragraph(str(header), make_header_style(bg_color)))

    table_data = [header_row]

    # Prepare data rows with Paragraph objects
    for row in data:
        row_data = []
        for idx, key in enumerate(column_keys):
            value = row.get(key, '')
            if isinstance(value, float):
                row_data.append(Paragraph(f"{value:,.2f}", cell_style_right))
            else:
                row_data.append(Paragraph(str(value) if value else '', cell_style))
        table_data.append(row_data)

    # Calculate column widths
    available_width = page_size[0] - 0.6*inch
    col_widths = [available_width / len(headers)] * len(headers)

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Build style commands with color coding for headers
    style_commands = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
    ]

    # Apply background colors to header cells
    for col_idx, bg_color in enumerate(header_bg_colors):
        style_commands.append(('BACKGROUND', (col_idx, 0), (col_idx, 0), colors.HexColor(bg_color)))

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============================================================================
# HR Report Data Extraction & Export Functions
# =============================================================================

def get_turnover_data(year=None):
    """Get turnover report data."""
    from django.utils import timezone
    from django.db.models.functions import TruncMonth

    if not year:
        year = timezone.now().year

    exits = Employee.objects.filter(
        status='SEPARATED',
        date_of_exit__year=year
    )

    by_month = exits.annotate(
        month=TruncMonth('date_of_exit')
    ).values('month').annotate(count=Count('id')).order_by('month')

    by_reason = exits.values('exit_reason').annotate(count=Count('id'))

    by_department = exits.values(
        department_name=F('department__name')
    ).annotate(count=Count('id')).order_by('-count')

    headers = ['Month', 'Exits', 'Exit Reason', 'Department']

    data = []
    for row in by_month:
        data.append({
            'month': row['month'].strftime('%B %Y') if row['month'] else '',
            'exits': row['count'],
            'exit_reason': '',
            'department': '',
        })
    for row in by_reason:
        data.append({
            'month': '',
            'exits': row['count'],
            'exit_reason': row['exit_reason'] or 'Unknown',
            'department': '',
        })
    for row in by_department:
        data.append({
            'month': '',
            'exits': row['count'],
            'exit_reason': '',
            'department': row['department_name'] or 'Unassigned',
        })

    return data, headers


def export_turnover(year=None, format='csv'):
    """Export turnover report with chart visualizations for PDF/Excel."""
    data, headers = get_turnover_data(year)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Turnover Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'turnover_{timestamp}.csv')

    # Extract chart data from the mixed flat data structure
    monthly_data = [{'name': r['month'], 'value': r['exits']} for r in data if r.get('month')]
    reason_data = [{'name': r['exit_reason'], 'value': r['exits']} for r in data if r.get('exit_reason')]
    dept_data = [{'name': r['department'], 'value': r['exits']} for r in data if r.get('department')]

    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'Turnover Data'},
        {'type': 'chart', 'chart_type': 'line', 'data': monthly_data,
         'title': 'Monthly Exit Trend', 'options': {}},
        {'type': 'chart', 'chart_type': 'pie', 'data': reason_data,
         'title': 'Exits by Reason', 'options': {'donut': True}},
        {'type': 'chart', 'chart_type': 'bar', 'data': dept_data[:10],
         'title': 'Top 10 Departments by Exits', 'options': {'horizontal': True}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'turnover_{timestamp}.pdf', title)
    else:
        return generate_excel_with_charts_response(sections, f'turnover_{timestamp}.xlsx', title)


def get_demographics_data():
    """Get demographics report data."""
    from django.utils import timezone

    queryset = Employee.objects.filter(status='ACTIVE')
    today = timezone.now().date()
    current_year = today.year

    headers = ['Category', 'Value', 'Count']

    data = []

    # By gender
    by_gender = queryset.values('gender').annotate(count=Count('id'))
    for row in by_gender:
        data.append({
            'category': 'Gender',
            'value': row['gender'] or 'Not Specified',
            'count': row['count'],
        })

    # By marital status
    by_marital = queryset.values('marital_status').annotate(count=Count('id'))
    for row in by_marital:
        data.append({
            'category': 'Marital Status',
            'value': row['marital_status'] or 'Not Specified',
            'count': row['count'],
        })

    # By age bracket
    age_brackets = {
        '18-25': 0, '26-30': 0, '31-35': 0, '36-40': 0,
        '41-45': 0, '46-50': 0, '51-55': 0, '56-60': 0, '60+': 0,
    }
    for dob in queryset.filter(date_of_birth__isnull=False).values_list('date_of_birth', flat=True):
        age = current_year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        if age <= 25:
            age_brackets['18-25'] += 1
        elif age <= 30:
            age_brackets['26-30'] += 1
        elif age <= 35:
            age_brackets['31-35'] += 1
        elif age <= 40:
            age_brackets['36-40'] += 1
        elif age <= 45:
            age_brackets['41-45'] += 1
        elif age <= 50:
            age_brackets['46-50'] += 1
        elif age <= 55:
            age_brackets['51-55'] += 1
        elif age <= 60:
            age_brackets['56-60'] += 1
        else:
            age_brackets['60+'] += 1

    for bracket, count in age_brackets.items():
        data.append({
            'category': 'Age Bracket',
            'value': bracket,
            'count': count,
        })

    return data, headers


def export_demographics(format='csv'):
    """Export demographics report with chart visualizations for PDF/Excel."""
    data, headers = get_demographics_data()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Demographics Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'demographics_{timestamp}.csv')

    # Extract chart data from the flat category-based structure
    gender_data = [{'name': r['value'], 'value': r['count']} for r in data if r.get('category') == 'Gender']
    marital_data = [{'name': r['value'], 'value': r['count']} for r in data if r.get('category') == 'Marital Status']
    age_data = [{'name': r['value'], 'value': r['count']} for r in data if r.get('category') == 'Age Bracket']

    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'Demographics Data'},
        {'type': 'chart', 'chart_type': 'pie', 'data': gender_data,
         'title': 'Gender Distribution', 'options': {'donut': True}},
        {'type': 'chart', 'chart_type': 'pie', 'data': marital_data,
         'title': 'Marital Status Distribution', 'options': {'donut': True}},
        {'type': 'chart', 'chart_type': 'area', 'data': age_data,
         'title': 'Age Distribution', 'options': {}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'demographics_{timestamp}.pdf', title)
    else:
        return generate_excel_with_charts_response(sections, f'demographics_{timestamp}.xlsx', title)


def get_leave_utilization_data(year=None):
    """Get leave utilization report data."""
    from django.utils import timezone
    from django.db.models.functions import TruncMonth

    if not year:
        year = timezone.now().year

    by_month = LeaveRequest.objects.filter(
        start_date__year=year,
        status='APPROVED'
    ).annotate(
        month=TruncMonth('start_date')
    ).values('month').annotate(
        count=Count('id'),
        total_days=Sum('number_of_days')
    ).order_by('month')

    by_leave_type = LeaveRequest.objects.filter(
        start_date__year=year,
        status='APPROVED'
    ).values(
        leave_type_name=F('leave_type__name')
    ).annotate(
        count=Count('id'),
        total_days=Sum('number_of_days')
    )

    by_department = LeaveRequest.objects.filter(
        start_date__year=year,
        status='APPROVED'
    ).values(
        department_name=F('employee__department__name')
    ).annotate(
        count=Count('id'),
        total_days=Sum('number_of_days')
    ).order_by('-total_days')

    headers = ['Section', 'Name', 'Requests', 'Total Days']

    data = []
    for row in by_month:
        data.append({
            'section': 'Monthly',
            'name': row['month'].strftime('%B %Y') if row['month'] else '',
            'requests': row['count'],
            'total_days': float(row['total_days'] or 0),
        })
    for row in by_leave_type:
        data.append({
            'section': 'Leave Type',
            'name': row['leave_type_name'] or 'Unknown',
            'requests': row['count'],
            'total_days': float(row['total_days'] or 0),
        })
    for row in by_department:
        data.append({
            'section': 'Department',
            'name': row['department_name'] or 'Unassigned',
            'requests': row['count'],
            'total_days': float(row['total_days'] or 0),
        })

    return data, headers


def export_leave_utilization(year=None, format='csv'):
    """Export leave utilization report with chart visualizations for PDF/Excel."""
    data, headers = get_leave_utilization_data(year)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Leave Utilization Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'leave_utilization_{timestamp}.csv')

    # Extract chart data from the section-based flat structure
    monthly_data = [{'name': r['name'], 'value': float(r['total_days'])}
                    for r in data if r.get('section') == 'Monthly']
    type_data = [{'name': r['name'], 'value': float(r['total_days'])}
                 for r in data if r.get('section') == 'Leave Type']
    dept_data = [{'name': r['name'], 'value': float(r['total_days'])}
                 for r in data if r.get('section') == 'Department']

    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'Leave Utilization Data'},
        {'type': 'chart', 'chart_type': 'line', 'data': monthly_data,
         'title': 'Monthly Leave Trend (Days)', 'options': {}},
        {'type': 'chart', 'chart_type': 'pie', 'data': type_data,
         'title': 'Leave Days by Type', 'options': {'donut': True}},
        {'type': 'chart', 'chart_type': 'bar', 'data': dept_data[:15],
         'title': 'Top 15 Departments by Leave Days', 'options': {'horizontal': True}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'leave_utilization_{timestamp}.pdf', title)
    else:
        return generate_excel_with_charts_response(sections, f'leave_utilization_{timestamp}.xlsx', title)


def get_employment_history_data(employee_id=None):
    """Get employment history report data for an employee."""
    queryset = EmploymentHistory.objects.select_related(
        'employee', 'previous_department', 'new_department',
        'previous_position', 'new_position',
        'previous_grade', 'new_grade',
    )

    if employee_id:
        queryset = queryset.filter(employee_id=employee_id)

    queryset = queryset.order_by('-effective_date')

    headers = [
        'Employee', 'Change Type', 'Effective Date',
        'Previous Department', 'New Department',
        'Previous Position', 'New Position',
        'Previous Grade', 'New Grade', 'Reason'
    ]

    data = []
    for record in queryset:
        data.append({
            'employee': record.employee.full_name if record.employee else '',
            'change_type': record.change_type or '',
            'effective_date': record.effective_date.strftime('%Y-%m-%d') if record.effective_date else '',
            'previous_department': record.previous_department.name if record.previous_department else '',
            'new_department': record.new_department.name if record.new_department else '',
            'previous_position': record.previous_position.title if record.previous_position else '',
            'new_position': record.new_position.title if record.new_position else '',
            'previous_grade': record.previous_grade.name if record.previous_grade else '',
            'new_grade': record.new_grade.name if record.new_grade else '',
            'reason': record.reason or '',
        })

    return data, headers


def export_employment_history(employee_id=None, format='csv'):
    """Export employment history report."""
    data, headers = get_employment_history_data(employee_id)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Employment History Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'employment_history_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'employment_history_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'employment_history_{timestamp}.csv')


def get_kpi_data():
    """Get KPI tracking report data."""
    headers = ['Category', 'KPI Name', 'Value']

    data = []

    # Performance KPIs
    total_employees = Employee.objects.filter(status='ACTIVE').count()
    completed = Appraisal.objects.filter(status__in=['COMPLETED', 'ACKNOWLEDGED']).count()
    total_appraisals = Appraisal.objects.count()
    completion_rate = (completed / total_appraisals * 100) if total_appraisals > 0 else 0
    avg_rating = Appraisal.objects.filter(
        overall_final_rating__isnull=False
    ).aggregate(avg=Avg('overall_final_rating'))['avg']

    data.append({'category': 'Performance', 'kpi_name': 'Total Employees', 'value': total_employees})
    data.append({'category': 'Performance', 'kpi_name': 'Completed Appraisals', 'value': completed})
    data.append({'category': 'Performance', 'kpi_name': 'Completion Rate (%)', 'value': round(completion_rate, 1)})
    data.append({'category': 'Performance', 'kpi_name': 'Average Rating', 'value': round(float(avg_rating), 1) if avg_rating else 'N/A'})

    # Training KPIs
    total_programs = TrainingProgram.objects.filter(is_active=True).count()
    total_sessions = TrainingSession.objects.count()
    completed_sessions = TrainingSession.objects.filter(status='COMPLETED').count()
    session_completion = (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0

    data.append({'category': 'Training', 'kpi_name': 'Active Programs', 'value': total_programs})
    data.append({'category': 'Training', 'kpi_name': 'Total Sessions', 'value': total_sessions})
    data.append({'category': 'Training', 'kpi_name': 'Completed Sessions', 'value': completed_sessions})
    data.append({'category': 'Training', 'kpi_name': 'Session Completion Rate (%)', 'value': round(session_completion, 1)})

    # Recruitment KPIs
    from django.utils import timezone
    current_year = timezone.now().year
    new_hires = Employee.objects.filter(date_of_joining__year=current_year).count()
    exits = Employee.objects.filter(status='SEPARATED', date_of_exit__year=current_year).count()
    active = Employee.objects.filter(status='ACTIVE').count()
    turnover_rate = (exits / active * 100) if active > 0 else 0

    data.append({'category': 'Recruitment', 'kpi_name': 'New Hires (This Year)', 'value': new_hires})
    data.append({'category': 'Recruitment', 'kpi_name': 'Exits (This Year)', 'value': exits})
    data.append({'category': 'Recruitment', 'kpi_name': 'Turnover Rate (%)', 'value': round(turnover_rate, 1)})

    return data, headers


def export_kpi_tracking(format='csv'):
    """Export KPI tracking report with chart visualizations for PDF/Excel."""
    data, headers = get_kpi_data()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "KPI Tracking Report"

    if format == 'csv':
        return generate_csv_response(data, headers, f'kpi_tracking_{timestamp}.csv')

    # Rating distribution bar chart - bucket appraisals by rating
    rating_buckets = {'Outstanding (4.5-5.0)': 0, 'Exceeds (3.5-4.5)': 0,
                      'Meets (2.5-3.5)': 0, 'Needs Improvement (1.5-2.5)': 0,
                      'Unsatisfactory (0-1.5)': 0}
    for rating in Appraisal.objects.filter(
        overall_final_rating__isnull=False
    ).values_list('overall_final_rating', flat=True):
        r = float(rating)
        if r >= 4.5:
            rating_buckets['Outstanding (4.5-5.0)'] += 1
        elif r >= 3.5:
            rating_buckets['Exceeds (3.5-4.5)'] += 1
        elif r >= 2.5:
            rating_buckets['Meets (2.5-3.5)'] += 1
        elif r >= 1.5:
            rating_buckets['Needs Improvement (1.5-2.5)'] += 1
        else:
            rating_buckets['Unsatisfactory (0-1.5)'] += 1

    rating_chart = [{'name': k, 'value': v} for k, v in rating_buckets.items() if v > 0]

    # Training programs by category for pie chart
    training_by_cat = [
        {'name': row['category'] or 'Uncategorized', 'value': row['count']}
        for row in TrainingProgram.objects.filter(is_active=True)
        .values('category').annotate(count=Count('id')).order_by('-count')
    ]

    sections = [
        {'type': 'table', 'data': data, 'headers': headers, 'subtitle': 'KPI Metrics'},
        {'type': 'chart', 'chart_type': 'bar', 'data': rating_chart,
         'title': 'Performance Rating Distribution', 'options': {}},
        {'type': 'chart', 'chart_type': 'pie', 'data': training_by_cat,
         'title': 'Active Training Programs by Category', 'options': {'donut': True}},
    ]

    if format == 'pdf':
        return generate_pdf_with_charts_response(sections, f'kpi_tracking_{timestamp}.pdf', title)
    else:
        return generate_excel_with_charts_response(sections, f'kpi_tracking_{timestamp}.xlsx', title)


def get_performance_appraisals_data(filters=None):
    """Get performance appraisals report data."""
    queryset = Appraisal.objects.select_related(
        'employee', 'employee__department', 'appraisal_cycle'
    ).order_by('-appraisal_cycle__year', 'employee__last_name')

    if filters:
        if filters.get('cycle'):
            queryset = queryset.filter(appraisal_cycle_id=filters['cycle'])
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        if filters.get('search'):
            search = filters['search']
            queryset = queryset.filter(
                Q(employee__first_name__icontains=search) |
                Q(employee__last_name__icontains=search) |
                Q(employee__employee_number__icontains=search)
            )

    headers = [
        'Employee', 'Employee Number', 'Department', 'Cycle',
        'Status', 'Self Rating', 'Manager Rating', 'Final Rating'
    ]

    data = []
    for a in queryset:
        data.append({
            'employee': a.employee.full_name if a.employee else '',
            'employee_number': a.employee.employee_number if a.employee else '',
            'department': a.employee.department.name if a.employee and a.employee.department else '',
            'cycle': a.appraisal_cycle.name if a.appraisal_cycle else '',
            'status': a.get_status_display() if hasattr(a, 'get_status_display') else a.status,
            'self_rating': float(a.overall_self_rating) if a.overall_self_rating else '',
            'manager_rating': float(a.overall_manager_rating) if a.overall_manager_rating else '',
            'final_rating': float(a.overall_final_rating) if a.overall_final_rating else '',
        })

    return data, headers


def export_performance_appraisals(filters=None, format='csv'):
    """Export performance appraisals report."""
    data, headers = get_performance_appraisals_data(filters)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Performance Appraisals Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'performance_appraisals_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'performance_appraisals_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'performance_appraisals_{timestamp}.csv')


def get_training_data():
    """Get training & development report data."""
    programs = TrainingProgram.objects.all().order_by('name')
    sessions = TrainingSession.objects.select_related('program').all()

    headers = [
        'Program Code', 'Program Name', 'Category', 'Type',
        'Duration Hours', 'Sessions', 'Mandatory'
    ]

    data = []
    for p in programs:
        session_count = sessions.filter(program=p).count()
        data.append({
            'program_code': p.code,
            'program_name': p.name,
            'category': p.get_category_display() if hasattr(p, 'get_category_display') else p.category,
            'type': p.get_training_type_display() if hasattr(p, 'get_training_type_display') else p.training_type,
            'duration_hours': float(p.duration_hours) if p.duration_hours else 0,
            'sessions': session_count,
            'mandatory': 'Yes' if p.is_mandatory else 'No',
        })

    return data, headers


def export_training(format='csv'):
    """Export training & development report."""
    data, headers = get_training_data()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    title = "Training & Development Report"

    if format == 'excel':
        return generate_excel_response(data, headers, f'training_{timestamp}.xlsx', title)
    elif format == 'pdf':
        return generate_pdf_response(data, headers, f'training_{timestamp}.pdf', title, landscape_mode=True)
    else:
        return generate_csv_response(data, headers, f'training_{timestamp}.csv')


class ReportExporter:
    """
    Utility class for exporting report data to various formats.
    Provides a simple interface for CSV, Excel, and PDF exports.
    """

    @staticmethod
    def export_data(data: list, headers: list, filename: str, format_type: str = 'csv', title: str = None):
        """
        Export data to the specified format.

        Args:
            data: List of dictionaries containing the data to export
            headers: List of column headers
            filename: Base filename (without extension)
            format_type: One of 'csv', 'excel', 'pdf'
            title: Optional title for the report

        Returns:
            HttpResponse with the exported file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Normalize data to use header keys
        normalized_data = []
        for row in data:
            normalized_row = {}
            for header in headers:
                # Try both original header and normalized key
                key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('%', '')
                normalized_row[key] = row.get(header, row.get(key, ''))
            normalized_data.append(normalized_row)

        if format_type.lower() == 'excel':
            return generate_excel_response(
                normalized_data, headers,
                f"{filename}_{timestamp}.xlsx",
                title=title
            )
        elif format_type.lower() == 'pdf':
            return generate_pdf_response(
                normalized_data, headers,
                f"{filename}_{timestamp}.pdf",
                title=title,
                landscape_mode=True
            )
        else:
            return generate_csv_response(
                normalized_data, headers,
                f"{filename}_{timestamp}.csv"
            )
