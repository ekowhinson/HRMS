"""
Payslip and Bank File generators supporting multiple formats (PDF, Excel, CSV/Text).
"""

import io
import csv
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional

from django.utils import timezone
from django.conf import settings
from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def get_org_settings():
    """Get organization settings from Django settings."""
    payroll_settings = getattr(settings, 'PAYROLL', {})
    return {
        'name': payroll_settings.get('ORGANIZATION_NAME', settings.HRMS_SETTINGS.get('ORGANIZATION_NAME', 'Your Organization')),
        'code': payroll_settings.get('ORGANIZATION_CODE', settings.HRMS_SETTINGS.get('ORGANIZATION_CODE', 'ORG')),
    }


class PayslipGenerator:
    """Generate payslips in multiple formats."""

    def __init__(self, payroll_item, period):
        self.item = payroll_item
        self.period = period
        self.emp = payroll_item.employee
        self.org = get_org_settings()

        # Collect allowances and deductions
        self._collect_components()
        self._calculate_ytd()

    def _collect_components(self):
        """Collect allowances and deductions from payroll item details."""
        self.allowances = []
        self.deductions = []
        self.employer_contributions = []
        self.loan_deductions = []
        self.arrear_allowances = []
        self.arrear_deductions = []

        for detail in self.item.details.all():
            comp = detail.pay_component
            amount = float(detail.amount)
            is_arrear = getattr(detail, 'is_arrear', False)

            if comp.component_type == 'EARNING' and comp.code != 'BASIC':
                target = self.arrear_allowances if is_arrear else self.allowances
                target.append({'name': comp.name, 'amount': amount})
            elif comp.component_type == 'DEDUCTION':
                target = self.arrear_deductions if is_arrear else self.deductions
                target.append({'name': comp.name, 'amount': amount})
                if not is_arrear and 'loan' in comp.name.lower():
                    self.loan_deductions.append({'name': comp.name, 'amount': amount})
            elif comp.component_type == 'EMPLOYER':
                self.employer_contributions.append({'name': comp.name, 'amount': amount})

        # Add statutory deductions
        if self.item.ssnit_employee and float(self.item.ssnit_employee) > 0:
            self.deductions.append({'name': 'Employee SSF', 'amount': float(self.item.ssnit_employee)})
        if self.item.paye and float(self.item.paye) > 0:
            self.deductions.append({'name': 'Income Tax', 'amount': float(self.item.paye)})

    def _calculate_ytd(self):
        """Calculate Year-to-Date values."""
        from .models import PayrollItem, PayrollItemDetail, PayrollRun

        current_year = self.period.year if self.period else timezone.now().year

        # Get all approved/paid payroll items for this employee in the current year
        ytd_items = PayrollItem.objects.filter(
            employee=self.emp,
            payroll_run__payroll_period__year=current_year,
            payroll_run__status__in=[PayrollRun.Status.COMPUTED, PayrollRun.Status.APPROVED, PayrollRun.Status.PAID]
        ).aggregate(
            total_earnings=Sum('gross_earnings'),
            total_ssf=Sum('ssnit_employee'),
            total_tax=Sum('paye'),
            total_net=Sum('net_salary'),
        )

        self.earnings_ytd = float(ytd_items['total_earnings'] or self.item.gross_earnings)
        self.ssf_ytd = float(ytd_items['total_ssf'] or self.item.ssnit_employee or 0)
        self.tax_ytd = float(ytd_items['total_tax'] or self.item.paye or 0)
        self.net_ytd = float(ytd_items['total_net'] or self.item.net_salary)

        # Calculate PF YTD
        pf_ytd = PayrollItemDetail.objects.filter(
            payroll_item__employee=self.emp,
            payroll_item__payroll_run__payroll_period__year=current_year,
            payroll_item__payroll_run__status__in=[PayrollRun.Status.COMPUTED, PayrollRun.Status.APPROVED, PayrollRun.Status.PAID],
            pay_component__code__icontains='PF',
            pay_component__component_type='DEDUCTION'
        ).aggregate(total=Sum('amount'))
        self.pf_ytd = float(pf_ytd['total'] or 0)

        # Calculate loan deductions YTD
        loan_ytd = PayrollItemDetail.objects.filter(
            payroll_item__employee=self.emp,
            payroll_item__payroll_run__payroll_period__year=current_year,
            payroll_item__payroll_run__status__in=[PayrollRun.Status.COMPUTED, PayrollRun.Status.APPROVED, PayrollRun.Status.PAID],
            pay_component__name__icontains='loan'
        ).aggregate(total=Sum('amount'))
        self.loans_ytd = float(loan_ytd['total'] or 0)

        # PF to date (all time)
        emp_pf_total = PayrollItemDetail.objects.filter(
            payroll_item__employee=self.emp,
            payroll_item__payroll_run__status__in=[PayrollRun.Status.COMPUTED, PayrollRun.Status.APPROVED, PayrollRun.Status.PAID],
            pay_component__code__icontains='PF',
            pay_component__component_type='DEDUCTION'
        ).aggregate(total=Sum('amount'))

        employer_pf_total = PayrollItemDetail.objects.filter(
            payroll_item__employee=self.emp,
            payroll_item__payroll_run__status__in=[PayrollRun.Status.COMPUTED, PayrollRun.Status.APPROVED, PayrollRun.Status.PAID],
            pay_component__code__icontains='PF',
            pay_component__component_type='EMPLOYER'
        ).aggregate(total=Sum('amount'))

        self.emp_pf_to_date = float(emp_pf_total['total'] or 0)
        self.employer_pf_to_date = float(employer_pf_total['total'] or 0)
        self.total_pf_to_date = self.emp_pf_to_date + self.employer_pf_to_date

    def _get_salary_notch_info(self):
        """Get salary notch information."""
        if self.emp.salary_notch:
            notch = self.emp.salary_notch
            if notch.level and notch.level.band:
                return f"Band {notch.level.band.code}/Level {notch.level.code}/Notch {notch.code}"
        return ''

    def generate_pdf(self) -> bytes:
        """Generate payslip as PDF."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        elements = []
        styles = getSampleStyleSheet()

        # Colors
        header_green = colors.HexColor('#008751')
        header_blue = colors.HexColor('#0077B6')

        # Styles
        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=16,
            textColor=header_green, alignment=TA_CENTER, spaceAfter=2, fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'], fontSize=12,
            alignment=TA_CENTER, spaceAfter=10, fontName='Helvetica-Bold'
        )
        logo_style = ParagraphStyle(
            'Logo', parent=styles['Normal'], fontSize=24,
            textColor=header_green, alignment=TA_LEFT, fontName='Helvetica-Bold'
        )
        # Cell styles with word wrapping
        cell_style = ParagraphStyle(
            'Cell', parent=styles['Normal'], fontSize=9,
            leading=11, wordWrap='CJK'
        )
        cell_style_bold = ParagraphStyle(
            'CellBold', parent=styles['Normal'], fontSize=9,
            leading=11, fontName='Helvetica-Bold', wordWrap='CJK'
        )
        cell_style_right = ParagraphStyle(
            'CellRight', parent=styles['Normal'], fontSize=9,
            leading=11, alignment=TA_RIGHT, wordWrap='CJK'
        )
        cell_style_white = ParagraphStyle(
            'CellWhite', parent=styles['Normal'], fontSize=9,
            leading=11, textColor=colors.white, fontName='Helvetica-Bold'
        )
        cell_style_white_right = ParagraphStyle(
            'CellWhiteRight', parent=styles['Normal'], fontSize=9,
            leading=11, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_RIGHT
        )

        period_name = self.period.name if self.period else 'N/A'
        notch_info = self._get_salary_notch_info()

        # Header
        header_data = [[
            Paragraph(f'<font color="#008751"><b>{self.org["code"]}</b></font>', logo_style),
            Paragraph(f'<u>{self.org["name"].upper()}</u>', title_style)
        ]]
        header_table = Table(header_data, colWidths=[3*cm, 13.5*cm])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(header_table)
        elements.append(Paragraph(f'<u>PAYSLIP FOR</u>&nbsp;&nbsp;&nbsp;&nbsp;<u>{period_name.upper()}</u>', subtitle_style))
        elements.append(Spacer(1, 10))

        # Employee Info - Use Paragraph for text wrapping
        emp_info_data = [
            [Paragraph('FULL NAME:', cell_style_bold), Paragraph(self.emp.full_name or '', cell_style),
             Paragraph('BANK NAME:', cell_style_bold), Paragraph(self.item.bank_name or '', cell_style)],
            [Paragraph('STAFF ID:', cell_style_bold), Paragraph(self.emp.employee_number or '', cell_style),
             Paragraph('BANK BRANCH:', cell_style_bold), Paragraph(self.item.bank_branch or '', cell_style)],
            [Paragraph('DEPARTMENT:', cell_style_bold), Paragraph(self.emp.department.name if self.emp.department else '', cell_style),
             Paragraph('ACCOUNT #:', cell_style_bold), Paragraph(self.item.bank_account_number or '', cell_style)],
            [Paragraph('LOCATION:', cell_style_bold), Paragraph(self.emp.work_location.name if self.emp.work_location else 'HEAD OFFICE', cell_style),
             Paragraph('SSN #:', cell_style_bold), Paragraph(self.emp.ssnit_number or '', cell_style)],
            [Paragraph('JOB TITLE:', cell_style_bold), Paragraph(self.emp.position.title if self.emp.position else '', cell_style),
             Paragraph('GRADE:', cell_style_bold), Paragraph(self.emp.grade.name if self.emp.grade else '', cell_style)],
            [Paragraph('OLD LEVEL/NOTCH:', cell_style_bold), Paragraph(notch_info, cell_style),
             Paragraph('NEW LEVEL/NOTCH:', cell_style_bold), Paragraph(notch_info, cell_style)],
        ]
        emp_table = Table(emp_info_data, colWidths=[3*cm, 5*cm, 3*cm, 5.5*cm])
        emp_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F5F5F5')),
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 10))

        # Basic Salary
        basic_data = [[
            Paragraph('BASIC SALARY', cell_style_white),
            Paragraph(f'{float(self.item.basic_salary):,.2f}', cell_style_white_right)
        ]]
        basic_table = Table(basic_data, colWidths=[13*cm, 3.5*cm])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), header_blue),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(basic_table)

        # Allowances and Deductions - Use Paragraph for text wrapping
        max_rows = max(len(self.allowances), len(self.deductions), 1)
        allow_ded_data = [[
            Paragraph('ALLOWANCES', cell_style_white),
            Paragraph('AMOUNT', cell_style_white_right),
            Paragraph('DEDUCTIONS', cell_style_white),
            Paragraph('AMOUNT', cell_style_white_right)
        ]]
        for i in range(max_rows):
            row = []
            if i < len(self.allowances):
                row.extend([
                    Paragraph(self.allowances[i]['name'], cell_style),
                    Paragraph(f"{self.allowances[i]['amount']:,.2f}", cell_style_right)
                ])
            else:
                row.extend([Paragraph('', cell_style), Paragraph('', cell_style)])
            if i < len(self.deductions):
                row.extend([
                    Paragraph(self.deductions[i]['name'], cell_style),
                    Paragraph(f"{self.deductions[i]['amount']:,.2f}", cell_style_right)
                ])
            else:
                row.extend([Paragraph('', cell_style), Paragraph('', cell_style)])
            allow_ded_data.append(row)

        allow_ded_table = Table(allow_ded_data, colWidths=[5.5*cm, 2.75*cm, 5.5*cm, 2.75*cm])
        allow_ded_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), header_blue),
            ('BACKGROUND', (2, 0), (3, 0), header_blue),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        elements.append(allow_ded_table)
        elements.append(Spacer(1, 10))

        # Backpay Arrears Section
        if self.arrear_allowances or self.arrear_deductions:
            header_orange = colors.HexColor('#E67E22')
            cell_style_orange_white = ParagraphStyle(
                'CellOrangeWhite', parent=styles['Normal'], fontSize=9,
                leading=11, textColor=colors.white, fontName='Helvetica-Bold'
            )
            cell_style_orange_white_right = ParagraphStyle(
                'CellOrangeWhiteRight', parent=styles['Normal'], fontSize=9,
                leading=11, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_RIGHT
            )

            arrear_max_rows = max(len(self.arrear_allowances), len(self.arrear_deductions), 1)
            arrear_data = [[
                Paragraph('BACKPAY ARREAR EARNINGS', cell_style_orange_white),
                Paragraph('AMOUNT', cell_style_orange_white_right),
                Paragraph('BACKPAY ARREAR DEDUCTIONS', cell_style_orange_white),
                Paragraph('AMOUNT', cell_style_orange_white_right)
            ]]
            for i in range(arrear_max_rows):
                row = []
                if i < len(self.arrear_allowances):
                    row.extend([
                        Paragraph(self.arrear_allowances[i]['name'], cell_style),
                        Paragraph(f"{self.arrear_allowances[i]['amount']:,.2f}", cell_style_right)
                    ])
                else:
                    row.extend([Paragraph('', cell_style), Paragraph('', cell_style)])
                if i < len(self.arrear_deductions):
                    row.extend([
                        Paragraph(self.arrear_deductions[i]['name'], cell_style),
                        Paragraph(f"{self.arrear_deductions[i]['amount']:,.2f}", cell_style_right)
                    ])
                else:
                    row.extend([Paragraph('', cell_style), Paragraph('', cell_style)])
                arrear_data.append(row)

            # Total arrears row
            arrear_earn_total = sum(a['amount'] for a in self.arrear_allowances)
            arrear_ded_total = sum(d['amount'] for d in self.arrear_deductions)
            arrear_data.append([
                Paragraph('TOTAL ARREAR EARNINGS', cell_style_bold),
                Paragraph(f'{arrear_earn_total:,.2f}', cell_style_right),
                Paragraph('TOTAL ARREAR DEDUCTIONS', cell_style_bold),
                Paragraph(f'{arrear_ded_total:,.2f}', cell_style_right),
            ])

            arrear_table = Table(arrear_data, colWidths=[5.5*cm, 2.75*cm, 5.5*cm, 2.75*cm])
            arrear_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), header_orange),
                ('BACKGROUND', (2, 0), (3, 0), header_orange),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF3E0')),
                ('LINEABOVE', (0, -1), (-1, -1), 1, header_orange),
            ]))
            elements.append(arrear_table)
            elements.append(Spacer(1, 10))

        # Pay Summary
        total_earnings = float(self.item.gross_earnings)
        total_deductions = float(self.item.total_deductions)
        net_salary = float(self.item.net_salary)

        summary_header_style = ParagraphStyle(
            'SummaryHeader', parent=styles['Normal'], fontSize=10,
            textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER
        )
        summary_header = [[Paragraph('PAY SUMMARY', summary_header_style)]]
        summary_header_table = Table(summary_header, colWidths=[16.5*cm])
        summary_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), header_blue),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_header_table)

        summary_data = [
            [Paragraph('TOTAL EARNINGS', cell_style_bold), Paragraph(f'{total_earnings:,.2f}', cell_style_right),
             Paragraph('TOTAL DEDUCTIONS', cell_style_bold), Paragraph(f'{total_deductions:,.2f}', cell_style_right)],
            [Paragraph('TAX RELIEF', cell_style_bold), Paragraph('0.00', cell_style_right),
             Paragraph('NET SALARY', cell_style_bold), Paragraph(f'{net_salary:,.2f}', cell_style_right)],
            [Paragraph('', cell_style), Paragraph('', cell_style), Paragraph('', cell_style), Paragraph('', cell_style)],
            [Paragraph('TOTAL EARNINGS YTD', cell_style_bold), Paragraph(f'{self.earnings_ytd:,.2f}', cell_style_right),
             Paragraph('EMPLOYEE PF TO DATE', cell_style_bold), Paragraph(f'{self.emp_pf_to_date:,.2f}', cell_style_right)],
            [Paragraph('EMPLOYEE SSF YTD', cell_style_bold), Paragraph(f'{self.ssf_ytd:,.2f}', cell_style_right),
             Paragraph('', cell_style), Paragraph('', cell_style)],
            [Paragraph('EMPLOYEE PF YTD', cell_style_bold), Paragraph(f'{self.pf_ytd:,.2f}', cell_style_right),
             Paragraph('EMPLOYER PF TO DATE', cell_style_bold), Paragraph(f'{self.employer_pf_to_date:,.2f}', cell_style_right)],
            [Paragraph('INCOME TAX YTD', cell_style_bold), Paragraph(f'{self.tax_ytd:,.2f}', cell_style_right),
             Paragraph('', cell_style), Paragraph('', cell_style)],
            [Paragraph('LOANS & ADV. YTD', cell_style_bold), Paragraph(f'{self.loans_ytd:,.2f}', cell_style_right),
             Paragraph('TOTAL PF TO DATE', cell_style_bold), Paragraph(f'{self.total_pf_to_date:,.2f}', cell_style_right)],
            [Paragraph('NET SALARY YTD', cell_style_bold), Paragraph(f'{self.net_ytd:,.2f}', cell_style_right),
             Paragraph('', cell_style), Paragraph('', cell_style)],
        ]

        summary_table = Table(summary_data, colWidths=[4.5*cm, 3.75*cm, 4.5*cm, 3.75*cm])
        summary_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('LINEBELOW', (3, 1), (3, 1), 1.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 10))

        # Loan Details
        loan_header = [[Paragraph('LOAN DETAILS', summary_header_style)]]
        loan_header_table = Table(loan_header, colWidths=[16.5*cm])
        loan_header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), header_blue),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(loan_header_table)

        if self.loan_deductions:
            loan_data = [
                [Paragraph(l['name'], cell_style), Paragraph(f"{l['amount']:,.2f}", cell_style_right)]
                for l in self.loan_deductions
            ]
            loan_table = Table(loan_data, colWidths=[12*cm, 4.5*cm])
            loan_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ]))
            elements.append(loan_table)
        else:
            empty_loan_style = ParagraphStyle(
                'EmptyLoan', parent=styles['Normal'], fontSize=9,
                textColor=colors.grey, alignment=TA_CENTER
            )
            empty_loan = Table([[Paragraph('No active loans', empty_loan_style)]], colWidths=[16.5*cm])
            empty_loan.setStyle(TableStyle([
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
            ]))
            elements.append(empty_loan)

        elements.append(Spacer(1, 20))

        # Footer
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        elements.append(Paragraph(f'Printed On: {timezone.now().isoformat()}', footer_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()

    def generate_excel(self) -> bytes:
        """Generate payslip as Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"

        # Styles
        header_font = Font(bold=True, size=14, color="008751")
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        money_font = Font(size=10)
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="0077B6", end_color="0077B6", fill_type="solid")

        period_name = self.period.name if self.period else 'N/A'
        notch_info = self._get_salary_notch_info()

        row = 1

        # Header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=self.org['name'].upper()).font = header_font
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"PAYSLIP FOR {period_name.upper()}").font = Font(bold=True, size=12)
        row += 2

        # Employee Info
        info_data = [
            ('FULL NAME:', self.emp.full_name, 'BANK NAME:', self.item.bank_name or ''),
            ('STAFF ID:', self.emp.employee_number, 'BANK BRANCH:', self.item.bank_branch or ''),
            ('DEPARTMENT:', self.emp.department.name if self.emp.department else '', 'ACCOUNT #:', self.item.bank_account_number or ''),
            ('LOCATION:', self.emp.work_location.name if self.emp.work_location else 'HEAD OFFICE', 'SOCIAL SECURITY #:', self.emp.ssnit_number or ''),
            ('JOB TITLE:', self.emp.position.title if self.emp.position else '', 'GRADE:', self.emp.grade.name if self.emp.grade else ''),
            ('LEVEL/NOTCH:', notch_info, '', ''),
        ]

        for data in info_data:
            ws.cell(row=row, column=1, value=data[0]).font = label_font
            ws.cell(row=row, column=2, value=data[1]).font = value_font
            ws.cell(row=row, column=4, value=data[2]).font = label_font
            ws.cell(row=row, column=5, value=data[3]).font = value_font
            row += 1

        row += 1

        # Basic Salary
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value="BASIC SALARY").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=4, value=float(self.item.basic_salary)).font = money_font
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        row += 2

        # Allowances Header
        ws.cell(row=row, column=1, value="ALLOWANCES").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=2, value="AMOUNT").font = section_font
        ws.cell(row=row, column=2).fill = section_fill
        ws.cell(row=row, column=4, value="DEDUCTIONS").font = section_font
        ws.cell(row=row, column=4).fill = section_fill
        ws.cell(row=row, column=5, value="AMOUNT").font = section_font
        ws.cell(row=row, column=5).fill = section_fill
        row += 1

        # Allowances and Deductions
        max_rows = max(len(self.allowances), len(self.deductions), 1)
        for i in range(max_rows):
            if i < len(self.allowances):
                ws.cell(row=row, column=1, value=self.allowances[i]['name'])
                ws.cell(row=row, column=2, value=self.allowances[i]['amount']).number_format = '#,##0.00'
            if i < len(self.deductions):
                ws.cell(row=row, column=4, value=self.deductions[i]['name'])
                ws.cell(row=row, column=5, value=self.deductions[i]['amount']).number_format = '#,##0.00'
            row += 1

        row += 1

        # Backpay Arrears
        if self.arrear_allowances or self.arrear_deductions:
            arrear_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
            arrear_font = Font(bold=True, size=11, color="FFFFFF")

            ws.cell(row=row, column=1, value="BACKPAY ARREAR EARNINGS").font = arrear_font
            ws.cell(row=row, column=1).fill = arrear_fill
            ws.cell(row=row, column=2, value="AMOUNT").font = arrear_font
            ws.cell(row=row, column=2).fill = arrear_fill
            ws.cell(row=row, column=4, value="BACKPAY ARREAR DEDUCTIONS").font = arrear_font
            ws.cell(row=row, column=4).fill = arrear_fill
            ws.cell(row=row, column=5, value="AMOUNT").font = arrear_font
            ws.cell(row=row, column=5).fill = arrear_fill
            row += 1

            arrear_max = max(len(self.arrear_allowances), len(self.arrear_deductions), 1)
            for i in range(arrear_max):
                if i < len(self.arrear_allowances):
                    ws.cell(row=row, column=1, value=self.arrear_allowances[i]['name'])
                    ws.cell(row=row, column=2, value=self.arrear_allowances[i]['amount']).number_format = '#,##0.00'
                if i < len(self.arrear_deductions):
                    ws.cell(row=row, column=4, value=self.arrear_deductions[i]['name'])
                    ws.cell(row=row, column=5, value=self.arrear_deductions[i]['amount']).number_format = '#,##0.00'
                row += 1

            # Arrear totals
            arrear_earn_total = sum(a['amount'] for a in self.arrear_allowances)
            arrear_ded_total = sum(d['amount'] for d in self.arrear_deductions)
            ws.cell(row=row, column=1, value="TOTAL ARREAR EARNINGS").font = label_font
            ws.cell(row=row, column=2, value=arrear_earn_total).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value="TOTAL ARREAR DEDUCTIONS").font = label_font
            ws.cell(row=row, column=5, value=arrear_ded_total).number_format = '#,##0.00'
            row += 2

        # Pay Summary
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="PAY SUMMARY").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1

        summary = [
            ('TOTAL EARNINGS', float(self.item.gross_earnings), 'TOTAL DEDUCTIONS', float(self.item.total_deductions)),
            ('TAX RELIEF', 0.00, 'NET SALARY', float(self.item.net_salary)),
            ('', '', '', ''),
            ('TOTAL EARNINGS YTD', self.earnings_ytd, 'EMPLOYEE PF TO DATE', self.emp_pf_to_date),
            ('EMPLOYEE SSF YTD', self.ssf_ytd, '', ''),
            ('EMPLOYEE PF YTD', self.pf_ytd, 'EMPLOYER PF TO DATE', self.employer_pf_to_date),
            ('INCOME TAX YTD', self.tax_ytd, '', ''),
            ('LOANS & ADV. DED. YTD', self.loans_ytd, 'TOTAL PF TO DATE', self.total_pf_to_date),
            ('NET SALARY YTD', self.net_ytd, '', ''),
        ]

        for s in summary:
            ws.cell(row=row, column=1, value=s[0]).font = label_font
            if s[1] != '':
                ws.cell(row=row, column=2, value=s[1]).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=s[2]).font = label_font
            if s[3] != '':
                ws.cell(row=row, column=5, value=s[3]).number_format = '#,##0.00'
            row += 1

        row += 1

        # Loan Details
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="LOAN DETAILS").font = section_font
        ws.cell(row=row, column=1).fill = section_fill
        row += 1

        if self.loan_deductions:
            for loan in self.loan_deductions:
                ws.cell(row=row, column=1, value=loan['name'])
                ws.cell(row=row, column=2, value=loan['amount']).number_format = '#,##0.00'
                row += 1
        else:
            ws.cell(row=row, column=1, value="No active loans")
            row += 1

        row += 2
        ws.cell(row=row, column=1, value=f"Printed On: {timezone.now().isoformat()}")

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_text(self) -> bytes:
        """Generate payslip as text/CSV file."""
        lines = []
        period_name = self.period.name if self.period else 'N/A'
        notch_info = self._get_salary_notch_info()

        lines.append('=' * 70)
        lines.append(self.org['name'].upper())
        lines.append(f"PAYSLIP FOR {period_name.upper()}")
        lines.append('=' * 70)
        lines.append('')

        # Employee Info
        lines.append('-' * 70)
        lines.append('EMPLOYEE INFORMATION')
        lines.append('-' * 70)
        lines.append(f"{'Full Name:':<25} {self.emp.full_name}")
        lines.append(f"{'Staff ID:':<25} {self.emp.employee_number}")
        lines.append(f"{'Department:':<25} {self.emp.department.name if self.emp.department else ''}")
        lines.append(f"{'Location:':<25} {self.emp.work_location.name if self.emp.work_location else 'HEAD OFFICE'}")
        lines.append(f"{'Job Title:':<25} {self.emp.position.title if self.emp.position else ''}")
        lines.append(f"{'Grade:':<25} {self.emp.grade.name if self.emp.grade else ''}")
        lines.append(f"{'Level/Notch:':<25} {notch_info}")
        lines.append(f"{'Bank Name:':<25} {self.item.bank_name or ''}")
        lines.append(f"{'Bank Branch:':<25} {self.item.bank_branch or ''}")
        lines.append(f"{'Account #:':<25} {self.item.bank_account_number or ''}")
        lines.append(f"{'SSNIT #:':<25} {self.emp.ssnit_number or ''}")
        lines.append('')

        # Basic Salary
        lines.append('-' * 70)
        lines.append(f"{'BASIC SALARY':<50} {float(self.item.basic_salary):>15,.2f}")
        lines.append('-' * 70)
        lines.append('')

        # Allowances
        lines.append('ALLOWANCES')
        lines.append('-' * 35)
        for a in self.allowances:
            lines.append(f"  {a['name']:<33} {a['amount']:>15,.2f}")
        lines.append('')

        # Deductions
        lines.append('DEDUCTIONS')
        lines.append('-' * 35)
        for d in self.deductions:
            lines.append(f"  {d['name']:<33} {d['amount']:>15,.2f}")
        lines.append('')

        # Backpay Arrears
        if self.arrear_allowances or self.arrear_deductions:
            lines.append('=' * 70)
            lines.append('BACKPAY ARREARS')
            lines.append('=' * 70)
            if self.arrear_allowances:
                lines.append('ARREAR EARNINGS')
                lines.append('-' * 35)
                for a in self.arrear_allowances:
                    lines.append(f"  {a['name']:<33} {a['amount']:>15,.2f}")
                arrear_earn_total = sum(a['amount'] for a in self.arrear_allowances)
                lines.append(f"  {'TOTAL ARREAR EARNINGS':<33} {arrear_earn_total:>15,.2f}")
                lines.append('')
            if self.arrear_deductions:
                lines.append('ARREAR DEDUCTIONS')
                lines.append('-' * 35)
                for d in self.arrear_deductions:
                    lines.append(f"  {d['name']:<33} {d['amount']:>15,.2f}")
                arrear_ded_total = sum(d['amount'] for d in self.arrear_deductions)
                lines.append(f"  {'TOTAL ARREAR DEDUCTIONS':<33} {arrear_ded_total:>15,.2f}")
                lines.append('')

        # Summary
        lines.append('=' * 70)
        lines.append('PAY SUMMARY')
        lines.append('=' * 70)
        lines.append(f"{'TOTAL EARNINGS:':<35} {float(self.item.gross_earnings):>15,.2f}")
        lines.append(f"{'TOTAL DEDUCTIONS:':<35} {float(self.item.total_deductions):>15,.2f}")
        lines.append(f"{'TAX RELIEF:':<35} {'0.00':>15}")
        lines.append('-' * 70)
        lines.append(f"{'NET SALARY:':<35} {float(self.item.net_salary):>15,.2f}")
        lines.append('=' * 70)
        lines.append('')

        # YTD
        lines.append('YEAR-TO-DATE SUMMARY')
        lines.append('-' * 35)
        lines.append(f"{'Total Earnings YTD:':<35} {self.earnings_ytd:>15,.2f}")
        lines.append(f"{'Employee SSF YTD:':<35} {self.ssf_ytd:>15,.2f}")
        lines.append(f"{'Employee PF YTD:':<35} {self.pf_ytd:>15,.2f}")
        lines.append(f"{'Income Tax YTD:':<35} {self.tax_ytd:>15,.2f}")
        lines.append(f"{'Loans & Adv. Ded. YTD:':<35} {self.loans_ytd:>15,.2f}")
        lines.append(f"{'Net Salary YTD:':<35} {self.net_ytd:>15,.2f}")
        lines.append('')
        lines.append(f"{'Employee PF to Date:':<35} {self.emp_pf_to_date:>15,.2f}")
        lines.append(f"{'Employer PF to Date:':<35} {self.employer_pf_to_date:>15,.2f}")
        lines.append(f"{'Total PF to Date:':<35} {self.total_pf_to_date:>15,.2f}")
        lines.append('')

        # Loan Details
        lines.append('LOAN DETAILS')
        lines.append('-' * 35)
        if self.loan_deductions:
            for loan in self.loan_deductions:
                lines.append(f"  {loan['name']:<33} {loan['amount']:>15,.2f}")
        else:
            lines.append('  No active loans')
        lines.append('')

        lines.append('=' * 70)
        lines.append(f"Printed On: {timezone.now().isoformat()}")
        lines.append('=' * 70)

        return '\n'.join(lines).encode('utf-8')


class BankFileGenerator:
    """Generate bank payment files in multiple formats."""

    def __init__(self, payroll_run, payroll_items):
        self.payroll_run = payroll_run
        self.items = payroll_items
        self.org = get_org_settings()

    def generate_csv(self) -> bytes:
        """Generate bank file as CSV."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            'Bank', 'Branch', 'Account Number', 'Account Name',
            'Employee Number', 'Net Salary', 'Reference'
        ])

        total_amount = 0
        for item in self.items:
            writer.writerow([
                item.bank_name or '',
                item.bank_branch or '',
                item.bank_account_number or '',
                item.employee.full_name,
                item.employee.employee_number,
                float(item.net_salary),
                item.payment_reference or f'{self.payroll_run.run_number}-{item.employee.employee_number}'
            ])
            total_amount += float(item.net_salary)

        # Summary
        writer.writerow([])
        writer.writerow(['Total Records:', len(self.items), 'Total Amount:', f'{total_amount:,.2f}'])

        output.seek(0)
        return output.getvalue().encode('utf-8')

    def generate_excel(self) -> bytes:
        """Generate bank file as Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Advice"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0077B6", end_color="0077B6", fill_type="solid")
        money_format = '#,##0.00'

        # Title
        ws.merge_cells('A1:G1')
        ws.cell(row=1, column=1, value=f"{self.org['name']} - BANK ADVICE").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Payroll Run: {self.payroll_run.run_number}")
        ws.cell(row=2, column=4, value=f"Period: {self.payroll_run.payroll_period.name if self.payroll_run.payroll_period else 'N/A'}")

        # Headers
        headers = ['Bank', 'Branch', 'Account Number', 'Account Name', 'Employee Number', 'Net Salary', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        total_amount = 0
        row = 5
        for item in self.items:
            ws.cell(row=row, column=1, value=item.bank_name or '')
            ws.cell(row=row, column=2, value=item.bank_branch or '')
            ws.cell(row=row, column=3, value=item.bank_account_number or '')
            ws.cell(row=row, column=4, value=item.employee.full_name)
            ws.cell(row=row, column=5, value=item.employee.employee_number)
            ws.cell(row=row, column=6, value=float(item.net_salary)).number_format = money_format
            ws.cell(row=row, column=7, value=item.payment_reference or f'{self.payroll_run.run_number}-{item.employee.employee_number}')
            total_amount += float(item.net_salary)
            row += 1

        # Summary
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=5, value=len(self.items)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_amount).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = money_format

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_pdf(self) -> bytes:
        """Generate bank file as PDF."""
        from reportlab.lib.pagesizes import A4, landscape

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )

        elements = []
        styles = getSampleStyleSheet()

        header_blue = colors.HexColor('#0077B6')

        # Title
        title_style = ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=14,
            alignment=TA_CENTER, spaceAfter=10
        )
        elements.append(Paragraph(f"{self.org['name']} - BANK ADVICE", title_style))

        subtitle_style = ParagraphStyle(
            'Subtitle', parent=styles['Normal'], fontSize=10,
            alignment=TA_CENTER, spaceAfter=20
        )
        period_name = self.payroll_run.payroll_period.name if self.payroll_run.payroll_period else 'N/A'
        elements.append(Paragraph(f"Payroll Run: {self.payroll_run.run_number} | Period: {period_name}", subtitle_style))

        # Cell styles for wrapping
        cell_style = ParagraphStyle(
            'BankCell', parent=styles['Normal'], fontSize=8,
            leading=10, wordWrap='CJK'
        )
        cell_style_right = ParagraphStyle(
            'BankCellRight', parent=styles['Normal'], fontSize=8,
            leading=10, alignment=TA_RIGHT, wordWrap='CJK'
        )
        cell_style_bold = ParagraphStyle(
            'BankCellBold', parent=styles['Normal'], fontSize=8,
            leading=10, fontName='Helvetica-Bold', wordWrap='CJK'
        )
        header_cell_style = ParagraphStyle(
            'BankHeaderCell', parent=styles['Normal'], fontSize=8,
            leading=10, textColor=colors.white, fontName='Helvetica-Bold'
        )
        header_cell_style_right = ParagraphStyle(
            'BankHeaderCellRight', parent=styles['Normal'], fontSize=8,
            leading=10, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_RIGHT
        )

        # Table header
        table_data = [[
            Paragraph('Bank', header_cell_style),
            Paragraph('Branch', header_cell_style),
            Paragraph('Account #', header_cell_style),
            Paragraph('Account Name', header_cell_style),
            Paragraph('Emp #', header_cell_style),
            Paragraph('Net Salary', header_cell_style_right),
            Paragraph('Reference', header_cell_style)
        ]]

        total_amount = 0
        for item in self.items:
            table_data.append([
                Paragraph(item.bank_name or '', cell_style),
                Paragraph(item.bank_branch or '', cell_style),
                Paragraph(item.bank_account_number or '', cell_style),
                Paragraph(item.employee.full_name, cell_style),
                Paragraph(item.employee.employee_number, cell_style),
                Paragraph(f'{float(item.net_salary):,.2f}', cell_style_right),
                Paragraph(item.payment_reference or f'{self.payroll_run.run_number}-{item.employee.employee_number}', cell_style)
            ])
            total_amount += float(item.net_salary)

        # Total row
        table_data.append([
            Paragraph('TOTAL', cell_style_bold),
            Paragraph('', cell_style),
            Paragraph('', cell_style),
            Paragraph('', cell_style),
            Paragraph(str(len(self.items)), cell_style_bold),
            Paragraph(f'{total_amount:,.2f}', cell_style_right),
            Paragraph('', cell_style)
        ])

        table = Table(table_data, colWidths=[3.5*cm, 4*cm, 3*cm, 5*cm, 2.5*cm, 3*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_blue),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
        ]))
        elements.append(table)

        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        elements.append(Paragraph(f'Generated On: {timezone.now().isoformat()}', footer_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read()
