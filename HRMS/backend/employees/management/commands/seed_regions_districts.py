"""
Management command to seed Region/District tables and update Employee records
from three Excel files containing region and district data.
"""

import os
import re

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from core.models import Country, Region, District
from employees.models import Employee


# Paths to Excel files (relative to project root)
EXCEL_FILES = [
    'district_staff_data.xlsx',
    'non_managers_headoffice_regionaloffice.xlsx',
    'managers_data.xlsx',
]


def generate_code(name, existing_codes, max_len):
    """Generate a unique uppercase code from a name, handling collisions."""
    # Strip non-alpha, take first max_len uppercase letters
    letters = re.sub(r'[^A-Za-z]', '', name).upper()
    base = letters[:max_len] if letters else 'UNK'
    code = base
    suffix = 1
    while code in existing_codes:
        code = f"{base[:max_len - len(str(suffix))]}{suffix}"
        suffix += 1
    existing_codes.add(code)
    return code


class Command(BaseCommand):
    help = 'Seed Region and District tables from Excel files and update employee residential data.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview changes without writing to database.',
        )

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        # __file__ = .../expene-tracker-ai/HRMS/backend/employees/management/commands/seed_regions_districts.py
        # 6 dirname calls -> .../expene-tracker-ai
        project_root = os.path.abspath(__file__)
        for _ in range(6):
            project_root = os.path.dirname(project_root)

        # ------------------------------------------------------------------
        # 1. Read all Excel files and combine rows
        # ------------------------------------------------------------------
        rows = []  # list of (employee_number_str, region_name, district_name)
        for fname in EXCEL_FILES:
            fpath = os.path.join(project_root, fname)
            if not os.path.exists(fpath):
                self.stderr.write(self.style.WARNING(f'File not found, skipping: {fpath}'))
                continue

            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb.active

            # Find column indices from header row
            headers = {cell.value.strip() if cell.value else '': idx
                       for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}

            emp_col = headers.get('EMPLOYEE_NUMBER')
            reg_col = headers.get('REGION')
            dist_col = headers.get('DISTRICT')

            if emp_col is None or reg_col is None or dist_col is None:
                self.stderr.write(self.style.ERROR(
                    f'{fname}: Missing required columns (EMPLOYEE_NUMBER, REGION, DISTRICT). '
                    f'Found: {list(headers.keys())}'
                ))
                wb.close()
                continue

            file_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                emp_num = row[emp_col]
                region = row[reg_col]
                district = row[dist_col]

                # Skip rows with missing key data
                if emp_num is None or region is None or district is None:
                    continue

                emp_str = str(int(emp_num)) if isinstance(emp_num, float) else str(emp_num).strip()
                region_name = str(region).strip()
                district_name = str(district).strip()

                if not emp_str or not region_name or not district_name:
                    continue

                rows.append((emp_str, region_name, district_name))
                file_count += 1

            wb.close()
            self.stdout.write(f'  Read {file_count} rows from {fname}')

        self.stdout.write(f'\nTotal rows combined: {len(rows)}')

        # De-duplicate by employee number (keep first occurrence)
        seen = set()
        unique_rows = []
        for emp_str, region_name, district_name in rows:
            if emp_str not in seen:
                seen.add(emp_str)
                unique_rows.append((emp_str, region_name, district_name))
        self.stdout.write(f'Unique employees in Excel: {len(unique_rows)}')

        # ------------------------------------------------------------------
        # 2. Collect unique regions and region-district pairs
        # ------------------------------------------------------------------
        unique_regions = set()
        unique_pairs = set()  # (region_name, district_name)
        for _, region_name, district_name in unique_rows:
            unique_regions.add(region_name)
            unique_pairs.add((region_name, district_name))

        self.stdout.write(f'Unique regions: {len(unique_regions)}')
        self.stdout.write(f'Unique region-district pairs: {len(unique_pairs)}')

        if dry_run:
            self.stdout.write(self.style.WARNING('\n-- DRY RUN: No changes will be made --'))
            self.stdout.write(f'\nRegions to create: {sorted(unique_regions)}')
            self.stdout.write(f'\nDistricts to create ({len(unique_pairs)}):')
            for reg, dist in sorted(unique_pairs):
                self.stdout.write(f'  {reg} -> {dist}')
            return

        # ------------------------------------------------------------------
        # 3. Ensure Country(GHA) exists
        # ------------------------------------------------------------------
        country, created = Country.objects.get_or_create(
            code='GHA',
            defaults={'name': 'Ghana', 'phone_code': '+233', 'currency_code': 'GHS'},
        )
        if created:
            self.stdout.write(self.style.SUCCESS('Created Country: Ghana (GHA)'))
        else:
            self.stdout.write('Country Ghana (GHA) already exists')

        # ------------------------------------------------------------------
        # 4. Create Region records
        # ------------------------------------------------------------------
        region_code_set = set(Region.objects.values_list('code', flat=True))
        region_map = {}  # region_name -> Region instance
        regions_created = 0

        # Pre-load existing regions by name (case-insensitive)
        for r in Region.objects.all():
            region_map[r.name.upper()] = r

        for region_name in sorted(unique_regions):
            key = region_name.upper()
            if key in region_map:
                continue
            code = generate_code(region_name, region_code_set, 3)
            region = Region.objects.create(
                code=code,
                name=region_name,
                country=country,
            )
            region_map[key] = region
            regions_created += 1

        self.stdout.write(self.style.SUCCESS(f'Regions created: {regions_created}'))
        self.stdout.write(f'Total regions in DB: {Region.objects.count()}')

        # ------------------------------------------------------------------
        # 5. Create District records
        # ------------------------------------------------------------------
        district_code_set = set(District.objects.values_list('code', flat=True))
        district_map = {}  # (region_name_upper, district_name_upper) -> District
        districts_created = 0

        # Pre-load existing districts
        for d in District.objects.select_related('region').all():
            district_map[(d.region.name.upper(), d.name.upper())] = d

        for region_name, district_name in sorted(unique_pairs):
            rkey = region_name.upper()
            dkey = (rkey, district_name.upper())
            if dkey in district_map:
                continue
            region_obj = region_map[rkey]
            code = generate_code(district_name, district_code_set, 5)
            district = District.objects.create(
                code=code,
                name=district_name,
                region=region_obj,
            )
            district_map[dkey] = district
            districts_created += 1

        self.stdout.write(self.style.SUCCESS(f'Districts created: {districts_created}'))
        self.stdout.write(f'Total districts in DB: {District.objects.count()}')

        # ------------------------------------------------------------------
        # 6. Update employee records
        # ------------------------------------------------------------------
        # Build employee lookup
        emp_lookup = {}
        for emp in Employee.objects.only('id', 'employee_number', 'residential_region', 'residential_district'):
            emp_lookup[emp.employee_number] = emp

        employees_updated = 0
        employees_skipped = 0
        employees_not_found = 0
        employees_to_update = []

        for emp_str, region_name, district_name in unique_rows:
            emp = emp_lookup.get(emp_str)
            if emp is None:
                employees_not_found += 1
                continue

            region_obj = region_map.get(region_name.upper())
            district_obj = district_map.get((region_name.upper(), district_name.upper()))

            if region_obj is None or district_obj is None:
                employees_skipped += 1
                continue

            # Skip if already set to same values
            if emp.residential_region_id == region_obj.id and emp.residential_district_id == district_obj.id:
                employees_skipped += 1
                continue

            emp.residential_region = region_obj
            emp.residential_district = district_obj
            employees_to_update.append(emp)
            employees_updated += 1

        # Bulk update in batches of 500
        if employees_to_update:
            with transaction.atomic():
                batch_size = 500
                for i in range(0, len(employees_to_update), batch_size):
                    batch = employees_to_update[i:i + batch_size]
                    Employee.objects.bulk_update(batch, ['residential_region', 'residential_district'], batch_size=batch_size)

        self.stdout.write(self.style.SUCCESS(f'\nEmployees updated: {employees_updated}'))
        self.stdout.write(f'Employees skipped (already set or missing data): {employees_skipped}')
        self.stdout.write(f'Employees not found in DB: {employees_not_found}')
        self.stdout.write(f'Employees with region set: {Employee.objects.filter(residential_region__isnull=False).count()}')
        self.stdout.write(self.style.SUCCESS('\nDone!'))
