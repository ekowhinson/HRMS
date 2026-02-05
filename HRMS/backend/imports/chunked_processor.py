"""
Chunked import processor for handling large files (100K+ rows).
Processes data in memory-efficient chunks using generators.
"""

import csv
import io
import logging
from typing import Iterator, Dict, List, Any, Optional, Generator
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from django.db import transaction, connection
from django.utils import timezone
from django.core.cache import cache

logger = logging.getLogger(__name__)


@dataclass
class ChunkResult:
    """Result of processing a single chunk."""
    success_count: int = 0
    error_count: int = 0
    skip_count: int = 0
    errors: List[Dict] = field(default_factory=list)


@dataclass
class ChunkedImportResult:
    """Final result of chunked import."""
    success_count: int = 0
    error_count: int = 0
    skip_count: int = 0
    errors: List[Dict] = field(default_factory=list)
    chunks_processed: int = 0
    total_chunks: int = 0
    processing_time_seconds: float = 0


class ChunkedFileReader:
    """
    Memory-efficient file reader that yields data in chunks.
    Supports CSV, Excel, and other formats.
    """

    def __init__(self, chunk_size: int = 1000):
        self.chunk_size = chunk_size

    def read_csv_chunks(
        self,
        content: bytes,
        encoding: str = None
    ) -> Generator[tuple, None, None]:
        """
        Read CSV file in chunks, yielding (headers, chunk_rows) tuples.
        First yield includes headers, subsequent yields have headers=None.
        """
        import chardet

        # Detect encoding if not provided
        if not encoding:
            result = chardet.detect(content[:10000])
            encoding = result.get('encoding', 'utf-8') or 'utf-8'

        try:
            text = content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = content.decode('latin-1')

        # Detect delimiter
        sample = text[:4096]
        delimiter = self._detect_delimiter(sample)

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)

        headers = None
        chunk = []
        row_count = 0

        for row in reader:
            if headers is None:
                headers = row
                yield (headers, [])
                continue

            chunk.append(row)
            row_count += 1

            if len(chunk) >= self.chunk_size:
                yield (None, chunk)
                chunk = []

        # Yield remaining rows
        if chunk:
            yield (None, chunk)

    def read_excel_chunks(
        self,
        content: bytes,
        sheet_name: str = None
    ) -> Generator[tuple, None, None]:
        """
        Read Excel file in chunks using openpyxl's read_only mode.
        Yields (headers, chunk_rows) tuples.
        """
        from openpyxl import load_workbook

        wb = load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True
        )

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        headers = None
        chunk = []

        for row in ws.iter_rows(values_only=True):
            # Convert to list of strings
            row_data = [
                str(cell) if cell is not None else ''
                for cell in row
            ]

            # Skip empty rows
            if not any(cell.strip() for cell in row_data):
                continue

            if headers is None:
                headers = row_data
                yield (headers, [])
                continue

            chunk.append(row_data)

            if len(chunk) >= self.chunk_size:
                yield (None, chunk)
                chunk = []

        # Yield remaining rows
        if chunk:
            yield (None, chunk)

        wb.close()

    def read_chunks(
        self,
        content: bytes,
        filename: str
    ) -> Generator[tuple, None, None]:
        """
        Auto-detect file type and read in chunks.
        """
        ext = filename.rsplit('.', 1)[-1].lower() if filename else 'csv'

        if ext in ('xlsx', 'xls'):
            yield from self.read_excel_chunks(content)
        else:
            yield from self.read_csv_chunks(content)

    def count_rows(self, content: bytes, filename: str) -> int:
        """
        Count total rows in file without loading all into memory.
        """
        ext = filename.rsplit('.', 1)[-1].lower() if filename else 'csv'

        if ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True
            )
            ws = wb.active
            count = sum(1 for _ in ws.iter_rows(values_only=True)) - 1  # Minus header
            wb.close()
            return max(0, count)
        else:
            # For CSV, count lines minus header
            import chardet
            result = chardet.detect(content[:10000])
            encoding = result.get('encoding', 'utf-8') or 'utf-8'
            try:
                text = content.decode(encoding)
            except:
                text = content.decode('latin-1')
            return max(0, text.count('\n') - 1)

    def _detect_delimiter(self, sample: str) -> str:
        """Detect CSV delimiter from sample."""
        delimiters = [',', '\t', ';', '|']
        counts = {d: sample.count(d) for d in delimiters}
        return max(counts, key=counts.get) if max(counts.values()) > 0 else ','


class ChunkedImportProcessor:
    """
    Process imports in chunks for memory efficiency.
    Handles files up to millions of rows.
    """

    DEFAULT_CHUNK_SIZE = 1000
    BATCH_SIZE = 500  # DB batch insert size

    def __init__(self, chunk_size: int = None):
        self.chunk_size = chunk_size or self.DEFAULT_CHUNK_SIZE
        self.reader = ChunkedFileReader(chunk_size=self.chunk_size)
        self._fk_cache = {}

    def process(self, job) -> ChunkedImportResult:
        """
        Process an import job in chunks.
        """
        from .models import ImportJob

        result = ChunkedImportResult()
        start_time = datetime.now()

        try:
            job.status = ImportJob.Status.IMPORTING
            job.started_at = timezone.now()
            job.save(update_fields=['status', 'started_at'])

            # Count total rows for progress tracking
            total_rows = self.reader.count_rows(job.file_data, job.original_filename)
            result.total_chunks = (total_rows // self.chunk_size) + 1

            # Pre-load FK caches
            self._cache_foreign_keys(job.target_model)

            # Get processor function
            processor = self._get_processor(job.target_model)
            if not processor:
                raise ValueError(f"Unknown target model: {job.target_model}")

            # Build header index from mapping
            header_index = None
            headers = None
            chunk_num = 0

            # Process file in chunks
            for chunk_headers, chunk_rows in self.reader.read_chunks(
                job.file_data,
                job.original_filename
            ):
                if chunk_headers is not None:
                    headers = chunk_headers
                    header_index = {h: i for i, h in enumerate(headers)}
                    continue

                if not chunk_rows:
                    continue

                chunk_num += 1
                result.chunks_processed = chunk_num

                # Process this chunk
                chunk_result = processor(
                    chunk_rows,
                    header_index,
                    job.column_mapping,
                    job
                )

                # Aggregate results
                result.success_count += chunk_result.success_count
                result.error_count += chunk_result.error_count
                result.skip_count += chunk_result.skip_count

                # Keep only first 100 errors
                if len(result.errors) < 100:
                    result.errors.extend(chunk_result.errors[:100 - len(result.errors)])

                # Update progress
                self._update_progress(
                    job.id,
                    result.success_count + result.error_count + result.skip_count,
                    total_rows,
                    chunk_num,
                    result.total_chunks
                )

                # Clear Django query cache periodically to prevent memory buildup
                if chunk_num % 10 == 0:
                    from django import db
                    db.reset_queries()

            # Calculate processing time
            result.processing_time_seconds = (datetime.now() - start_time).total_seconds()

            # Update job
            job.success_count = result.success_count
            job.error_count = result.error_count
            job.skip_count = result.skip_count
            job.errors = result.errors
            job.status = ImportJob.Status.COMPLETED
            job.completed_at = timezone.now()
            job.save()

        except Exception as e:
            logger.exception(f"Chunked import job {job.id} failed: {str(e)}")
            job.fail(str(e))
            result.errors.append({
                'type': 'system',
                'message': str(e)
            })

        return result

    def _get_processor(self, target_model: str):
        """Get the processor function for a target model."""
        processors = {
            'employees': self._process_employees_chunk,
            'departments': self._process_departments_chunk,
            'positions': self._process_positions_chunk,
            'grades': self._process_grades_chunk,
            'divisions': self._process_divisions_chunk,
            'directorates': self._process_directorates_chunk,
            'banks': self._process_banks_chunk,
            'bank_branches': self._process_bank_branches_chunk,
            'bank_accounts': self._process_bank_accounts_chunk,
            'transactions': self._process_transactions_chunk,
            'leave_balances': self._process_leave_balances_chunk,
            'leave_types': self._process_leave_types_chunk,
            'staff_categories': self._process_staff_categories_chunk,
            'salary_bands': self._process_salary_bands_chunk,
            'salary_levels': self._process_salary_levels_chunk,
            'salary_notches': self._process_salary_notches_chunk,
            'pay_components': self._process_pay_components_chunk,
            'work_locations': self._process_work_locations_chunk,
        }
        return processors.get(target_model)

    def _cache_foreign_keys(self, target_model: str):
        """Pre-load and cache FK lookups for efficiency."""
        from organization.models import Department, JobPosition, JobGrade, Division, Directorate, WorkLocation
        from leave.models import LeaveType
        from payroll.models import PayComponent, Bank, BankBranch, StaffCategory, SalaryBand, SalaryLevel, SalaryNotch
        from employees.models import Employee

        self._fk_cache = {
            'departments': {},
            'positions': {},
            'grades': {},
            'leave_types': {},
            'pay_components': {},
            'employees': {},
            'divisions': {},
            'directorates': {},
            'work_locations': {},
            'banks': {},
            'bank_branches': {},
            'staff_categories': {},
            'salary_bands': {},
            'salary_levels': {},
            'salary_notches': {},
        }

        # Only load caches relevant to target model to save memory
        if target_model in ('employees', 'bank_accounts', 'leave_balances', 'transactions'):
            # Cache employees by employee_number (use iterator for large tables)
            for e in Employee.objects.only('id', 'employee_number').iterator(chunk_size=5000):
                self._fk_cache['employees'][e.employee_number.lower()] = e.id

        if target_model in ('employees',):
            for d in Department.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['departments'][d.code.lower()] = d.id
                self._fk_cache['departments'][d.name.lower()] = d.id

            for p in JobPosition.objects.only('id', 'code', 'title').iterator():
                self._fk_cache['positions'][p.code.lower()] = p.id
                self._fk_cache['positions'][p.title.lower()] = p.id

            for g in JobGrade.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['grades'][g.code.lower()] = g.id
                self._fk_cache['grades'][g.name.lower()] = g.id

            for div in Division.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['divisions'][div.code.lower()] = div.id
                self._fk_cache['divisions'][div.name.lower()] = div.id

            for d in Directorate.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['directorates'][d.code.lower()] = d.id
                self._fk_cache['directorates'][d.name.lower()] = d.id

            for wl in WorkLocation.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['work_locations'][wl.code.lower()] = wl.id
                self._fk_cache['work_locations'][wl.name.lower()] = wl.id

            for sc in StaffCategory.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['staff_categories'][sc.code.lower()] = sc.id
                self._fk_cache['staff_categories'][sc.name.lower()] = sc.id

            for sn in SalaryNotch.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['salary_notches'][sn.code.lower()] = sn.id
                self._fk_cache['salary_notches'][sn.name.lower()] = sn.id

        if target_model in ('bank_accounts',):
            for b in Bank.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['banks'][b.code.lower()] = b.id
                self._fk_cache['banks'][b.name.lower()] = b.id

            for bb in BankBranch.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['bank_branches'][bb.code.lower()] = bb.id
                self._fk_cache['bank_branches'][bb.name.lower()] = bb.id

        if target_model in ('transactions',):
            for pc in PayComponent.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['pay_components'][pc.code.lower()] = pc.id
                self._fk_cache['pay_components'][pc.name.lower()] = pc.id

        if target_model in ('leave_balances',):
            for lt in LeaveType.objects.only('id', 'code', 'name').iterator():
                self._fk_cache['leave_types'][lt.code.lower()] = lt.id
                self._fk_cache['leave_types'][lt.name.lower()] = lt.id

    def _resolve_fk_id(self, cache_key: str, value: str) -> Optional[int]:
        """Resolve a foreign key value to its ID from cache."""
        if not value:
            return None
        value_str = str(value).lower().strip()
        return self._fk_cache.get(cache_key, {}).get(value_str)

    def _extract_row_data(
        self,
        row: List,
        header_index: Dict[str, int],
        mapping: Dict[str, str]
    ) -> Dict[str, Any]:
        """Extract mapped data from a row."""
        data = {}
        for source_col, target_field in mapping.items():
            if source_col in header_index:
                idx = header_index[source_col]
                if idx < len(row):
                    value = row[idx]
                    if value is not None:
                        value = str(value).strip()
                        if value:
                            data[target_field] = value
        return data

    def _parse_date(self, value: str):
        """Parse date string to date object."""
        if not value:
            return None

        from datetime import datetime as dt

        value_str = str(value).strip()

        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
            '%d/%m/%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
        ]

        for fmt in formats:
            try:
                return dt.strptime(value_str, fmt).date()
            except ValueError:
                continue

        try:
            from dateutil.parser import parse as parse_date
            return parse_date(value_str).date()
        except:
            pass

        return None

    def _parse_decimal(self, value: str) -> Optional[Decimal]:
        """Parse string to Decimal."""
        if not value:
            return None
        try:
            # Remove currency symbols and commas
            cleaned = str(value).replace(',', '').replace('GH₵', '').replace('$', '').strip()
            return Decimal(cleaned)
        except:
            return None

    def _update_progress(
        self,
        job_id: str,
        processed: int,
        total: int,
        chunk_num: int,
        total_chunks: int
    ):
        """Update progress in cache for real-time tracking."""
        cache.set(
            f'import_progress_{job_id}',
            {
                'processed': processed,
                'total': total,
                'percentage': round((processed / total * 100) if total > 0 else 0, 1),
                'chunk': chunk_num,
                'total_chunks': total_chunks,
            },
            timeout=3600
        )

    # ===== Chunk Processors =====

    def _process_employees_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of employee records."""
        from employees.models import Employee

        result = ChunkResult()
        employees_to_create = []
        employees_to_update = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)

                if not data:
                    result.skip_count += 1
                    continue

                # Get employee number
                emp_number = data.get('employee_number', '').strip()
                if not emp_number:
                    result.skip_count += 1
                    continue

                # Check if exists
                existing_id = self._fk_cache['employees'].get(emp_number.lower())

                # Prepare employee data
                emp_data = {
                    'employee_number': emp_number,
                    'first_name': data.get('first_name', ''),
                    'last_name': data.get('last_name', ''),
                    'middle_name': data.get('middle_name', ''),
                    'gender': self._normalize_gender(data.get('gender', '')),
                    'date_of_birth': self._parse_date(data.get('date_of_birth')),
                    'email': data.get('email', ''),
                    'phone_number': data.get('phone_number', ''),
                    'ghana_card_number': data.get('ghana_card_number', ''),
                    'ssnit_number': data.get('ssnit_number', ''),
                    'tin_number': data.get('tin_number', ''),
                    'date_of_joining': self._parse_date(data.get('date_of_joining')) or timezone.now().date(),
                }

                # Resolve FKs
                if 'department' in data:
                    emp_data['department_id'] = self._resolve_fk_id('departments', data['department'])
                if 'position' in data:
                    emp_data['position_id'] = self._resolve_fk_id('positions', data['position'])
                if 'grade' in data:
                    emp_data['grade_id'] = self._resolve_fk_id('grades', data['grade'])
                if 'division' in data:
                    emp_data['division_id'] = self._resolve_fk_id('divisions', data['division'])
                if 'directorate' in data:
                    emp_data['directorate_id'] = self._resolve_fk_id('directorates', data['directorate'])
                if 'work_location' in data:
                    emp_data['work_location_id'] = self._resolve_fk_id('work_locations', data['work_location'])
                if 'staff_category' in data:
                    emp_data['staff_category_id'] = self._resolve_fk_id('staff_categories', data['staff_category'])
                if 'salary_notch' in data:
                    emp_data['salary_notch_id'] = self._resolve_fk_id('salary_notches', data['salary_notch'])

                if existing_id:
                    emp_data['id'] = existing_id
                    employees_to_update.append(emp_data)
                else:
                    employees_to_create.append(emp_data)
                    # Add to cache for deduplication within chunk
                    self._fk_cache['employees'][emp_number.lower()] = True

            except Exception as e:
                result.error_count += 1
                result.errors.append({
                    'row': row_num + 1,
                    'error': str(e)
                })

        # Bulk create new employees
        if employees_to_create:
            with transaction.atomic():
                created = Employee.objects.bulk_create(
                    [Employee(**data) for data in employees_to_create],
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        # Bulk update existing employees
        if employees_to_update:
            with transaction.atomic():
                for emp_data in employees_to_update:
                    emp_id = emp_data.pop('id')
                    Employee.objects.filter(id=emp_id).update(**emp_data)
                    result.success_count += 1

        return result

    def _normalize_gender(self, value: str) -> str:
        """Normalize gender value."""
        if not value:
            return 'O'
        value = value.lower().strip()
        if value in ('m', 'male', 'man'):
            return 'M'
        elif value in ('f', 'female', 'woman'):
            return 'F'
        return 'O'

    def _process_departments_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of department records."""
        from organization.models import Department

        result = ChunkResult()
        departments_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                # Check if exists
                if code.lower() in self._fk_cache.get('departments', {}):
                    result.skip_count += 1
                    continue

                departments_to_create.append(Department(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if departments_to_create:
            with transaction.atomic():
                created = Department.objects.bulk_create(
                    departments_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_positions_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of position records."""
        from organization.models import JobPosition

        result = ChunkResult()
        positions_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                title = data.get('title', '').strip()

                if not code and not title:
                    result.skip_count += 1
                    continue

                if code.lower() in self._fk_cache.get('positions', {}):
                    result.skip_count += 1
                    continue

                positions_to_create.append(JobPosition(
                    code=code or title[:20].upper().replace(' ', '_'),
                    title=title or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if positions_to_create:
            with transaction.atomic():
                created = JobPosition.objects.bulk_create(
                    positions_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_grades_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of grade records."""
        from organization.models import JobGrade

        result = ChunkResult()
        grades_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                if code.lower() in self._fk_cache.get('grades', {}):
                    result.skip_count += 1
                    continue

                grades_to_create.append(JobGrade(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    level=int(data.get('level', 1)),
                    min_salary=self._parse_decimal(data.get('min_salary')) or Decimal('0'),
                    max_salary=self._parse_decimal(data.get('max_salary')) or Decimal('0'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if grades_to_create:
            with transaction.atomic():
                created = JobGrade.objects.bulk_create(
                    grades_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_divisions_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of division records."""
        from organization.models import Division

        result = ChunkResult()
        divisions_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                divisions_to_create.append(Division(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if divisions_to_create:
            with transaction.atomic():
                created = Division.objects.bulk_create(
                    divisions_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_directorates_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of directorate records."""
        from organization.models import Directorate

        result = ChunkResult()
        directorates_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                directorates_to_create.append(Directorate(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if directorates_to_create:
            with transaction.atomic():
                created = Directorate.objects.bulk_create(
                    directorates_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_banks_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of bank records."""
        from payroll.models import Bank

        result = ChunkResult()
        banks_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                banks_to_create.append(Bank(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    swift_code=data.get('swift_code', ''),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if banks_to_create:
            with transaction.atomic():
                created = Bank.objects.bulk_create(
                    banks_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_bank_branches_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of bank branch records."""
        from payroll.models import BankBranch, Bank

        result = ChunkResult()
        branches_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()
                bank_ref = data.get('bank', '').strip()

                if not name:
                    result.skip_count += 1
                    continue

                bank_id = self._resolve_fk_id('banks', bank_ref)
                if not bank_id:
                    result.skip_count += 1
                    continue

                branches_to_create.append(BankBranch(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name,
                    bank_id=bank_id,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if branches_to_create:
            with transaction.atomic():
                created = BankBranch.objects.bulk_create(
                    branches_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_bank_accounts_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of bank account records."""
        from employees.models import BankAccount

        result = ChunkResult()
        accounts_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                emp_ref = data.get('employee', '').strip()
                account_number = data.get('account_number', '').strip()

                if not emp_ref or not account_number:
                    result.skip_count += 1
                    continue

                employee_id = self._resolve_fk_id('employees', emp_ref)
                if not employee_id:
                    result.skip_count += 1
                    continue

                bank_id = self._resolve_fk_id('banks', data.get('bank', ''))
                branch_id = self._resolve_fk_id('bank_branches', data.get('branch', ''))

                accounts_to_create.append(BankAccount(
                    employee_id=employee_id,
                    account_number=account_number,
                    account_name=data.get('account_name', ''),
                    bank_id=bank_id,
                    branch_id=branch_id,
                    is_primary=data.get('is_primary', '').lower() in ('yes', 'true', '1'),
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if accounts_to_create:
            with transaction.atomic():
                created = BankAccount.objects.bulk_create(
                    accounts_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_transactions_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of employee transaction records."""
        from payroll.models import EmployeeTransaction

        result = ChunkResult()
        transactions_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                emp_ref = data.get('employee', '').strip()
                component_ref = data.get('pay_component', '').strip()

                if not emp_ref or not component_ref:
                    result.skip_count += 1
                    continue

                employee_id = self._resolve_fk_id('employees', emp_ref)
                component_id = self._resolve_fk_id('pay_components', component_ref)

                if not employee_id or not component_id:
                    result.skip_count += 1
                    continue

                transactions_to_create.append(EmployeeTransaction(
                    employee_id=employee_id,
                    pay_component_id=component_id,
                    override_amount=self._parse_decimal(data.get('amount')),
                    effective_from=self._parse_date(data.get('effective_from')) or timezone.now().date(),
                    effective_to=self._parse_date(data.get('effective_to')),
                    is_recurring=data.get('is_recurring', '').lower() in ('yes', 'true', '1'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if transactions_to_create:
            with transaction.atomic():
                created = EmployeeTransaction.objects.bulk_create(
                    transactions_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_leave_balances_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of leave balance records."""
        from leave.models import LeaveBalance

        result = ChunkResult()
        balances_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                emp_ref = data.get('employee', '').strip()
                leave_type_ref = data.get('leave_type', '').strip()

                if not emp_ref or not leave_type_ref:
                    result.skip_count += 1
                    continue

                employee_id = self._resolve_fk_id('employees', emp_ref)
                leave_type_id = self._resolve_fk_id('leave_types', leave_type_ref)

                if not employee_id or not leave_type_id:
                    result.skip_count += 1
                    continue

                balances_to_create.append(LeaveBalance(
                    employee_id=employee_id,
                    leave_type_id=leave_type_id,
                    year=int(data.get('year', timezone.now().year)),
                    opening_balance=self._parse_decimal(data.get('opening_balance')) or Decimal('0'),
                    earned=self._parse_decimal(data.get('earned')) or Decimal('0'),
                    taken=self._parse_decimal(data.get('taken')) or Decimal('0'),
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if balances_to_create:
            with transaction.atomic():
                created = LeaveBalance.objects.bulk_create(
                    balances_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_leave_types_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of leave type records."""
        from leave.models import LeaveType

        result = ChunkResult()
        types_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                types_to_create.append(LeaveType(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    default_days=int(data.get('default_days', 0)),
                    is_paid=data.get('is_paid', '').lower() in ('yes', 'true', '1'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if types_to_create:
            with transaction.atomic():
                created = LeaveType.objects.bulk_create(
                    types_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_staff_categories_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of staff category records."""
        from payroll.models import StaffCategory

        result = ChunkResult()
        categories_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                categories_to_create.append(StaffCategory(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if categories_to_create:
            with transaction.atomic():
                created = StaffCategory.objects.bulk_create(
                    categories_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_salary_bands_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of salary band records."""
        from payroll.models import SalaryBand

        result = ChunkResult()
        bands_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                bands_to_create.append(SalaryBand(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if bands_to_create:
            with transaction.atomic():
                created = SalaryBand.objects.bulk_create(
                    bands_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_salary_levels_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of salary level records."""
        from payroll.models import SalaryLevel

        result = ChunkResult()
        levels_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                band_id = self._resolve_fk_id('salary_bands', data.get('band', ''))

                levels_to_create.append(SalaryLevel(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    band_id=band_id,
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if levels_to_create:
            with transaction.atomic():
                created = SalaryLevel.objects.bulk_create(
                    levels_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_salary_notches_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of salary notch records."""
        from payroll.models import SalaryNotch

        result = ChunkResult()
        notches_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                level_id = self._resolve_fk_id('salary_levels', data.get('level', ''))

                notches_to_create.append(SalaryNotch(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    level_id=level_id,
                    base_salary=self._parse_decimal(data.get('base_salary')) or Decimal('0'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if notches_to_create:
            with transaction.atomic():
                created = SalaryNotch.objects.bulk_create(
                    notches_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_pay_components_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of pay component records."""
        from payroll.models import PayComponent

        result = ChunkResult()
        components_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                comp_type = data.get('type', 'earning').lower()
                if comp_type in ('earning', 'earnings', 'e'):
                    comp_type = 'earning'
                elif comp_type in ('deduction', 'deductions', 'd'):
                    comp_type = 'deduction'

                components_to_create.append(PayComponent(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    component_type=comp_type,
                    is_taxable=data.get('is_taxable', '').lower() in ('yes', 'true', '1'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if components_to_create:
            with transaction.atomic():
                created = PayComponent.objects.bulk_create(
                    components_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result

    def _process_work_locations_chunk(
        self,
        rows: List[List],
        header_index: Dict[str, int],
        mapping: Dict[str, str],
        job
    ) -> ChunkResult:
        """Process a chunk of work location records."""
        from organization.models import WorkLocation

        result = ChunkResult()
        locations_to_create = []

        for row_num, row in enumerate(rows):
            try:
                data = self._extract_row_data(row, header_index, mapping)
                if not data:
                    result.skip_count += 1
                    continue

                code = data.get('code', '').strip()
                name = data.get('name', '').strip()

                if not code and not name:
                    result.skip_count += 1
                    continue

                locations_to_create.append(WorkLocation(
                    code=code or name[:20].upper().replace(' ', '_'),
                    name=name or code,
                    address=data.get('address', ''),
                    is_headquarters=data.get('is_headquarters', '').lower() in ('yes', 'true', '1'),
                    is_active=True
                ))

            except Exception as e:
                result.error_count += 1
                result.errors.append({'row': row_num + 1, 'error': str(e)})

        if locations_to_create:
            with transaction.atomic():
                created = WorkLocation.objects.bulk_create(
                    locations_to_create,
                    batch_size=self.BATCH_SIZE,
                    ignore_conflicts=True
                )
                result.success_count += len(created)

        return result


def process_import_chunked(job) -> ChunkedImportResult:
    """Convenience function to process an import job using chunked processor."""
    processor = ChunkedImportProcessor()
    return processor.process(job)
