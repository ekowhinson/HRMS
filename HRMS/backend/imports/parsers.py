"""
File parsers for different file types.
"""

import csv
import io
import chardet
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple
import logging

logger = logging.getLogger(__name__)


@dataclass
class ParsedData:
    """Result of parsing a file."""
    headers: List[str]
    rows: List[List[Any]]
    total_rows: int
    sample_rows: List[List[Any]]
    file_type: str
    encoding: Optional[str] = None
    sheet_name: Optional[str] = None
    metadata: Optional[Dict] = None


class FileParser:
    """Base class for file parsing."""

    SUPPORTED_TYPES = ['csv', 'xlsx', 'xls', 'txt', 'pdf']

    @staticmethod
    def detect_type(filename: str, content: bytes = None) -> str:
        """Detect file type from filename or content."""
        if filename:
            ext = filename.rsplit('.', 1)[-1].lower()
            if ext in FileParser.SUPPORTED_TYPES:
                return ext

        # Try to detect from content
        if content:
            # Check for Excel signature
            if content[:4] == b'PK\x03\x04':  # ZIP (xlsx)
                return 'xlsx'
            if content[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':  # OLE (xls)
                return 'xls'
            if content[:4] == b'%PDF':
                return 'pdf'

        return 'csv'  # Default to CSV

    @staticmethod
    def detect_encoding(content: bytes) -> str:
        """Detect text file encoding."""
        result = chardet.detect(content)
        return result.get('encoding', 'utf-8') or 'utf-8'

    @staticmethod
    def parse(file_type: str, content: bytes, filename: str = None) -> ParsedData:
        """
        Parse file content based on type.
        Returns headers, all rows, and sample data.
        """
        parsers = {
            'csv': CSVParser.parse,
            'txt': TextParser.parse,
            'xlsx': ExcelParser.parse,
            'xls': ExcelParser.parse,
            'pdf': PDFParser.parse,
        }

        parser = parsers.get(file_type)
        if not parser:
            raise ValueError(f"Unsupported file type: {file_type}")

        return parser(content, filename)


class CSVParser:
    """Parse CSV files with encoding detection."""

    @staticmethod
    def parse(content: bytes, filename: str = None) -> ParsedData:
        """Parse CSV content."""
        # Detect encoding
        encoding = FileParser.detect_encoding(content)

        try:
            text = content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            # Fallback to latin-1 which accepts all byte values
            text = content.decode('latin-1')
            encoding = 'latin-1'

        # Detect delimiter
        sample = text[:4096]
        delimiter = CSVParser._detect_delimiter(sample)

        # Parse CSV
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            return ParsedData(
                headers=[],
                rows=[],
                total_rows=0,
                sample_rows=[],
                file_type='csv',
                encoding=encoding
            )

        headers = rows[0]
        data_rows = rows[1:]

        return ParsedData(
            headers=headers,
            rows=data_rows,
            total_rows=len(data_rows),
            sample_rows=data_rows[:10],
            file_type='csv',
            encoding=encoding,
            metadata={'delimiter': delimiter}
        )

    @staticmethod
    def _detect_delimiter(sample: str) -> str:
        """Detect CSV delimiter from sample."""
        delimiters = [',', '\t', ';', '|']
        counts = {}

        for delim in delimiters:
            counts[delim] = sample.count(delim)

        # Return the delimiter with highest count
        return max(counts, key=counts.get) if max(counts.values()) > 0 else ','


class TextParser:
    """Parse delimited text files."""

    @staticmethod
    def parse(content: bytes, filename: str = None) -> ParsedData:
        """Parse text file (same as CSV but with different defaults)."""
        return CSVParser.parse(content, filename)


class ExcelParser:
    """Parse Excel files using openpyxl."""

    @staticmethod
    def parse(content: bytes, filename: str = None, sheet_name: str = None) -> ParsedData:
        """Parse Excel file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

        # Load workbook from bytes
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)

        # Get sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        # Read all data
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert row tuple to list, handling None values
            row_data = [
                str(cell) if cell is not None else ''
                for cell in row
            ]
            # Skip completely empty rows
            if any(cell.strip() for cell in row_data):
                rows.append(row_data)

        wb.close()

        if not rows:
            return ParsedData(
                headers=[],
                rows=[],
                total_rows=0,
                sample_rows=[],
                file_type='xlsx',
                sheet_name=sheet_name
            )

        headers = rows[0]
        data_rows = rows[1:]

        return ParsedData(
            headers=headers,
            rows=data_rows,
            total_rows=len(data_rows),
            sample_rows=data_rows[:10],
            file_type='xlsx',
            sheet_name=sheet_name,
            metadata={'sheets': wb.sheetnames if hasattr(wb, 'sheetnames') else [sheet_name]}
        )


class PDFParser:
    """Extract tables from PDF using pdfplumber."""

    @staticmethod
    def parse(content: bytes, filename: str = None) -> ParsedData:
        """Parse PDF file and extract tables."""
        try:
            import pdfplumber
        except ImportError:
            raise ImportError("pdfplumber is required for PDF parsing. Install with: pip install pdfplumber")

        all_rows = []
        headers = None

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue

                    for i, row in enumerate(table):
                        if row is None:
                            continue

                        # Clean row data
                        cleaned_row = [
                            str(cell).strip() if cell else ''
                            for cell in row
                        ]

                        # Skip empty rows
                        if not any(cleaned_row):
                            continue

                        # First non-empty row is headers
                        if headers is None:
                            headers = cleaned_row
                        else:
                            all_rows.append(cleaned_row)

        if not headers:
            return ParsedData(
                headers=[],
                rows=[],
                total_rows=0,
                sample_rows=[],
                file_type='pdf'
            )

        return ParsedData(
            headers=headers,
            rows=all_rows,
            total_rows=len(all_rows),
            sample_rows=all_rows[:10],
            file_type='pdf'
        )


def parse_file(content: bytes, filename: str) -> ParsedData:
    """
    Convenience function to parse a file.
    Automatically detects file type and parses.
    """
    file_type = FileParser.detect_type(filename, content)
    return FileParser.parse(file_type, content, filename)
