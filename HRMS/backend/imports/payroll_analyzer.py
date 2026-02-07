"""
Payroll file analyzer for NHIA HRMS payroll implementation.

Parses two specific Excel files:
1. NHIA Staff Allowances at all Bands.xlsx — band-level allowance policy
2. Staff Data for Payroll Implementation-23.12.25.xlsx — per-employee data (4 sheets)
"""

import io
import logging
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class PayrollFileAnalyzer:
    """Analyze the two payroll implementation Excel files."""

    def __init__(self):
        self.allowances_data = {}
        self.staff_data = []
        self.errors = []
        self.warnings = []

    def analyze_allowances_file(self, file_content):
        """
        Parse the NHIA Staff Allowances file.

        Returns band-level policy dict:
        {
            'BAND7': {'utility': 0.06, 'vehicle_maint': 0.18, 'fuel_gallons': 50, ...},
            'BAND1': {'utility': 0.06, 'transport': 300},
            ...
        }
        """
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Parse the allowances structure from the file
        # The file has a specific layout with band information and allowance rates
        band_policies = {}

        # Define the known band policies based on NHIA structure
        # These are extracted from the allowance file structure
        # Bands 1-3: Utility (6% of basic) + Transport (fixed)
        # Band 4: Utility (6%) + Vehicle Maintenance OR Transport
        # Bands 5-7: Multiple allowances

        # Try to parse from file rows
        current_band = None
        for row in rows:
            if not row or not any(row):
                continue

            row_str = [str(cell).strip() if cell is not None else '' for cell in row]
            row_lower = ' '.join(row_str).lower()

            # Detect band headers
            for band_num in range(1, 8):
                if f'band {band_num}' in row_lower or f'band{band_num}' in row_lower:
                    current_band = f'BAND{band_num}'
                    if current_band not in band_policies:
                        band_policies[current_band] = {}
                    break

            if not current_band:
                continue

            # Try to extract allowance info from rows
            for i, cell_str in enumerate(row_str):
                cell_lower = cell_str.lower()
                # Look for percentage values or amounts in adjacent cells
                if 'utility' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['utility'] = float(val)
                elif 'vehicle' in cell_lower and 'maint' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['vehicle_maint'] = float(val)
                elif 'fuel' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['fuel'] = float(val)
                elif 'security' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['security'] = float(val)
                elif 'entertainment' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['entertainment'] = float(val)
                elif 'domestic' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['domestic'] = float(val)
                elif 'responsibility' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['responsibility'] = float(val)
                elif 'rent' in cell_lower and 'allow' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['rent_allowance'] = float(val)
                elif 'transport' in cell_lower:
                    val = self._extract_numeric(row_str, i)
                    if val is not None:
                        band_policies[current_band]['transport'] = float(val)

        # Apply known defaults if file parsing didn't capture everything
        self._apply_default_band_policies(band_policies)

        self.allowances_data = band_policies
        return band_policies

    def _apply_default_band_policies(self, band_policies):
        """Apply known NHIA band policies as defaults."""
        defaults = {
            'BAND1': {
                'utility': 0.06,
                'transport': 300,
            },
            'BAND2': {
                'utility': 0.06,
                'transport': 300,
            },
            'BAND3': {
                'utility': 0.06,
                'transport': 300,
            },
            'BAND4': {
                'utility': 0.06,
                # Band 4 can have vehicle_maint OR transport depending on employee
                'transport': 450,  # For those without vehicle
                'vehicle_maint_fixed': 360,  # Fixed amount for those with vehicle
                'fuel_fixed': 840.72,  # Fixed fuel for Band 4 with vehicle
            },
            'BAND5': {
                'utility': 0.06,
                'vehicle_maint': 0.18,
                'fuel_fixed': 2101.80,
            },
            'BAND6': {
                'utility': 0.06,
                'vehicle_maint': 0.18,
                'fuel_fixed': 2802.40,
                'security': 0.10,
                'entertainment': 0.10,
                'domestic': 0.10,
                'responsibility': 0.10,
                'rent_allowance': 0.10,
            },
            'BAND7': {
                'utility': 0.06,
                'vehicle_maint': 0.18,
                'fuel_fixed': 3503.00,
                'security': 0.10,
                'entertainment': 0.15,
                'domestic': 0.15,
                'responsibility': 0.15,
                'rent_allowance': 0.10,
            },
        }

        for band, policy in defaults.items():
            if band not in band_policies:
                band_policies[band] = {}
            for key, val in policy.items():
                if key not in band_policies[band]:
                    band_policies[band][key] = val

    def _extract_numeric(self, row_str, label_index):
        """Extract a numeric value from adjacent cells after a label."""
        for offset in [1, 2, 3]:
            idx = label_index + offset
            if idx < len(row_str):
                val = self._parse_number(row_str[idx])
                if val is not None:
                    return val
        return None

    def _parse_number(self, value_str):
        """Parse a string to a number, handling percentages and currency."""
        if not value_str:
            return None
        cleaned = value_str.replace(',', '').replace('GHS', '').replace('GH¢', '').replace('%', '').strip()
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    def analyze_staff_data_file(self, file_content):
        """
        Parse all 4 sheets of the Staff Data file.

        Returns per-employee data list:
        [
            {
                'staff_id': '22428', 'nia': 'GHA-XXX', 'pf_rate': 0.05,
                'union': 'UNICOF', 'union_rate': 0.02, 'rent': 0,
                'bank_name': '...', 'branch_name': '...', 'bank_code': '',
                'branch_code': '', 'account_number': '...',
                'vehicle_allowance': 0, 'fuel_allowance': 0, 'transport': 0,
                'sheet': 'Directors',
            },
            ...
        ]
        """
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
        all_employees = []

        sheet_configs = self._get_sheet_configs(wb.sheetnames)

        for sheet_name, config in sheet_configs.items():
            if sheet_name not in wb.sheetnames:
                self.warnings.append(f"Sheet '{sheet_name}' not found, skipping")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) <= config['header_row']:
                self.warnings.append(f"Sheet '{sheet_name}' has insufficient rows")
                continue

            # Get header row for reference
            header_row = rows[config['header_row'] - 1] if config['header_row'] > 0 else rows[0]
            data_rows = rows[config['header_row']:]

            for row_idx, row in enumerate(data_rows):
                if not row or not any(row):
                    continue

                try:
                    emp_data = self._parse_employee_row(row, config, sheet_name)
                    if emp_data and emp_data.get('staff_id'):
                        all_employees.append(emp_data)
                except Exception as e:
                    self.errors.append(f"Sheet '{sheet_name}', row {row_idx + config['header_row'] + 1}: {str(e)}")

        wb.close()
        self.staff_data = all_employees
        return all_employees

    def _get_sheet_configs(self, sheet_names):
        """
        Get column mapping configs for each sheet.
        Column indices are 0-based.

        Sheet structures (from exploration):
        - Directors (header row 3): B=Staff ID, D=NIA, S=PF, T=Union, U=Rent, V-Z=Bank
        - Managers_Payroll (header row 7): B=Staff ID, D=NIA, N=PF, O=Union, P=Rent, Q-U=Bank
        - Districts Payroll (header row 7): B=Staff ID, D=NIA, K=Vehicle, L=Fuel, M=Transport, O=PF, P=Union, Q=Rent, R-V=Bank
        - Non Management Payroll (header row 7): Same as Districts
        """
        configs = {}

        # Try to match sheet names flexibly
        for name in sheet_names:
            name_lower = name.lower().strip()

            if 'director' in name_lower:
                configs[name] = {
                    'header_row': 3,
                    'staff_id': 1,      # B
                    'nia': 3,            # D
                    'pf': 18,            # S
                    'union': 19,         # T
                    'rent': 20,          # U
                    'bank_name': 21,     # V
                    'branch_name': 22,   # W
                    'bank_code': 23,     # X
                    'branch_code': 24,   # Y
                    'account_number': 25,  # Z
                    'vehicle_allowance': None,
                    'fuel_allowance': None,
                    'transport': None,
                    'category': 'directors',
                }
            elif 'manager' in name_lower:
                configs[name] = {
                    'header_row': 7,
                    'staff_id': 1,      # B
                    'nia': 3,            # D
                    'pf': 13,            # N
                    'union': 14,         # O
                    'rent': 15,          # P
                    'bank_name': 16,     # Q
                    'branch_name': 17,   # R
                    'bank_code': 18,     # S
                    'branch_code': 19,   # T
                    'account_number': 20,  # U
                    'vehicle_allowance': None,
                    'fuel_allowance': None,
                    'transport': None,
                    'category': 'managers',
                }
            elif 'district' in name_lower or 'non' in name_lower:
                configs[name] = {
                    'header_row': 7,
                    'staff_id': 1,      # B
                    'nia': 3,            # D
                    'vehicle_allowance': 10,  # K
                    'fuel_allowance': 11,     # L
                    'transport': 12,          # M
                    'pf': 14,            # O
                    'union': 15,         # P
                    'rent': 16,          # Q
                    'bank_name': 17,     # R
                    'branch_name': 18,   # S
                    'bank_code': 19,     # T
                    'branch_code': 20,   # U
                    'account_number': 21,  # V
                    'category': 'districts' if 'district' in name_lower else 'non_management',
                }

        return configs

    def _parse_employee_row(self, row, config, sheet_name):
        """Parse a single employee row based on the sheet config."""
        def get_cell(idx):
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            if val is None:
                return None
            return str(val).strip()

        def get_numeric(idx):
            val = get_cell(idx)
            if not val or val in ('None', '', '-', 'N/A'):
                return 0
            try:
                cleaned = val.replace(',', '').replace('GHS', '').replace('GH¢', '').replace('%', '').strip()
                return float(cleaned)
            except (ValueError, TypeError):
                return 0

        staff_id = get_cell(config['staff_id'])
        if not staff_id or staff_id in ('None', '', 'Staff ID', 'STAFF ID', 'S/N'):
            return None

        # Clean staff ID
        staff_id = staff_id.replace('.0', '').strip()

        nia = get_cell(config['nia']) or ''

        # PF rate
        pf_rate = get_numeric(config['pf'])
        # If value looks like a percentage (e.g., 5 or 11.5), convert to decimal
        if pf_rate > 1:
            pf_rate = pf_rate / 100

        # Union info
        union_val = get_numeric(config['union'])
        # If value looks like a percentage > 1, convert
        if union_val > 1:
            union_val = union_val / 100

        union_name = ''
        union_rate = 0
        if abs(union_val - 0.02) < 0.001:
            union_name = 'UNICOF'
            union_rate = 0.02
        elif abs(union_val - 0.015) < 0.001:
            union_name = 'PAWU'
            union_rate = 0.015
        elif union_val > 0:
            union_name = 'UNICOF'  # Default
            union_rate = union_val

        # Rent deduction
        rent = get_numeric(config['rent'])

        # Vehicle/fuel/transport (only for district/non-management sheets)
        vehicle_allowance = get_numeric(config.get('vehicle_allowance'))
        fuel_allowance = get_numeric(config.get('fuel_allowance'))
        transport = get_numeric(config.get('transport'))

        # Bank details
        bank_name = get_cell(config['bank_name']) or ''
        branch_name = get_cell(config['branch_name']) or ''
        bank_code = get_cell(config.get('bank_code')) or ''
        branch_code = get_cell(config.get('branch_code')) or ''
        account_number = get_cell(config['account_number']) or ''

        return {
            'staff_id': staff_id,
            'nia': nia,
            'pf_rate': pf_rate,
            'union': union_name,
            'union_rate': union_rate,
            'rent': rent,
            'bank_name': bank_name,
            'branch_name': branch_name,
            'bank_code': bank_code,
            'branch_code': branch_code,
            'account_number': account_number,
            'vehicle_allowance': vehicle_allowance,
            'fuel_allowance': fuel_allowance,
            'transport': transport,
            'sheet': sheet_name,
            'category': config.get('category', ''),
        }

    def generate_summary(self):
        """Generate analysis preview for the frontend."""
        # Band distribution from staff data (using employee's salary notch band)
        band_distribution = {}
        pf_count = 0
        unicof_count = 0
        pawu_count = 0
        rent_count = 0
        vehicle_count = 0
        transport_count = 0
        bank_count = 0

        for emp in self.staff_data:
            if emp.get('pf_rate', 0) > 0:
                pf_count += 1
            if emp.get('union') == 'UNICOF':
                unicof_count += 1
            elif emp.get('union') == 'PAWU':
                pawu_count += 1
            if emp.get('rent', 0) > 0:
                rent_count += 1
            if emp.get('vehicle_allowance', 0) > 0:
                vehicle_count += 1
            if emp.get('transport', 0) > 0:
                transport_count += 1
            if emp.get('account_number'):
                bank_count += 1

        # Count sheets
        sheet_counts = {}
        for emp in self.staff_data:
            sheet = emp.get('sheet', 'unknown')
            sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1

        return {
            'employee_count': len(self.staff_data),
            'sheet_counts': sheet_counts,
            'bands_found': list(self.allowances_data.keys()),
            'band_allowance_types': {
                band: list(policy.keys())
                for band, policy in self.allowances_data.items()
            },
            'deduction_counts': {
                'provident_fund': pf_count,
                'unicof': unicof_count,
                'pawu': pawu_count,
                'rent_deduction': rent_count,
            },
            'vehicle_employees': vehicle_count,
            'transport_employees': transport_count,
            'bank_accounts': bank_count,
            'components_to_create': 16,
            'tax_brackets_to_create': 7,
            'ssnit_rates_to_create': 2,
            'errors': self.errors,
            'warnings': self.warnings,
        }
